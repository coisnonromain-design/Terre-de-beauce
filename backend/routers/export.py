"""
Export routes (CSV/Excel)
"""
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime
import io
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from ..core.database import db

router = APIRouter(prefix="/export", tags=["Export"])


def create_excel_header_style():
    return {
        'fill': PatternFill(start_color="1A4D2E", end_color="1A4D2E", fill_type="solid"),
        'font': Font(bold=True, color="FFFFFF"),
        'alignment': Alignment(horizontal="center")
    }


@router.get("/factures")
async def export_factures(format: str = Query("csv", enum=["csv", "excel"]), statut: Optional[str] = None):
    """Export des factures en CSV ou Excel"""
    query = {}
    if statut:
        query["statut"] = statut
    
    factures = await db.factures.find(query, {"_id": 0}).sort("date_emission", -1).to_list(1000)
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Factures"
        
        style = create_excel_header_style()
        headers = ["N° Facture", "Date", "Client", "Chantier", "Montant HT", "TVA", "Montant TTC", "Statut", "Contrat CCPA"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = style['fill']
            cell.font = style['font']
            cell.alignment = style['alignment']
        
        for row, f in enumerate(factures, 2):
            ws.cell(row=row, column=1, value=f.get('numero', ''))
            ws.cell(row=row, column=2, value=f.get('date_emission', ''))
            ws.cell(row=row, column=3, value=f.get('client_raison_sociale', ''))
            ws.cell(row=row, column=4, value=f.get('chantier_reference', ''))
            ws.cell(row=row, column=5, value=f.get('montant_ht', 0))
            ws.cell(row=row, column=6, value=f.get('montant_tva', 0))
            ws.cell(row=row, column=7, value=f.get('montant_ttc', 0))
            ws.cell(row=row, column=8, value=f.get('statut', ''))
            ws.cell(row=row, column=9, value=f.get('contrat_numero', ''))
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=factures_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    else:
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(["N° Facture", "Date", "Client", "Chantier", "Montant HT", "TVA", "Montant TTC", "Statut", "Contrat CCPA"])
        
        for f in factures:
            writer.writerow([
                f.get('numero', ''),
                f.get('date_emission', ''),
                f.get('client_raison_sociale', ''),
                f.get('chantier_reference', ''),
                f.get('montant_ht', 0),
                f.get('montant_tva', 0),
                f.get('montant_ttc', 0),
                f.get('statut', ''),
                f.get('contrat_numero', '')
            ])
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=factures_{datetime.now().strftime('%Y%m%d')}.csv"}
        )


@router.get("/pointages")
async def export_pointages(format: str = Query("csv", enum=["csv", "excel"]), chauffeur_id: Optional[str] = None, chantier_id: Optional[str] = None):
    """Export des pointages en CSV ou Excel"""
    query = {}
    if chauffeur_id:
        query["chauffeur_id"] = chauffeur_id
    if chantier_id:
        query["chantier_id"] = chantier_id
    
    pointages = await db.pointages.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Pointages"
        
        style = create_excel_header_style()
        headers = ["Date", "Chauffeur", "Chantier", "Client", "Heures", "Volume Total", "Distance Total", "Nb Tours"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = style['fill']
            cell.font = style['font']
        
        for row, p in enumerate(pointages, 2):
            ws.cell(row=row, column=1, value=p.get('date', ''))
            ws.cell(row=row, column=2, value=p.get('chauffeur_nom', ''))
            ws.cell(row=row, column=3, value=p.get('chantier_reference', ''))
            ws.cell(row=row, column=4, value=p.get('client_nom', ''))
            ws.cell(row=row, column=5, value=p.get('heures_travaillees', 0))
            ws.cell(row=row, column=6, value=p.get('total_volume', 0))
            ws.cell(row=row, column=7, value=p.get('total_distance', 0))
            ws.cell(row=row, column=8, value=p.get('nombre_tours', 0))
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=pointages_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    else:
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(["Date", "Chauffeur", "Chantier", "Client", "Heures", "Volume Total", "Distance Total", "Nb Tours"])
        
        for p in pointages:
            writer.writerow([
                p.get('date', ''),
                p.get('chauffeur_nom', ''),
                p.get('chantier_reference', ''),
                p.get('client_nom', ''),
                p.get('heures_travaillees', 0),
                p.get('total_volume', 0),
                p.get('total_distance', 0),
                p.get('nombre_tours', 0)
            ])
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=pointages_{datetime.now().strftime('%Y%m%d')}.csv"}
        )


@router.get("/chantiers")
async def export_chantiers(format: str = Query("csv", enum=["csv", "excel"]), statut: Optional[str] = None):
    """Export des chantiers en CSV ou Excel"""
    query = {}
    if statut:
        query["statut"] = statut
    
    chantiers = await db.chantiers.find(query, {"_id": 0}).sort("date_debut", -1).to_list(1000)
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Chantiers"
        
        style = create_excel_header_style()
        headers = ["Référence", "Client", "Lieu", "Date Début", "Date Fin", "Type Transport", "Gasoil", "Statut"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = style['fill']
            cell.font = style['font']
        
        for row, c in enumerate(chantiers, 2):
            ws.cell(row=row, column=1, value=c.get('reference', ''))
            ws.cell(row=row, column=2, value=c.get('client_nom', ''))
            ws.cell(row=row, column=3, value=c.get('lieu', ''))
            ws.cell(row=row, column=4, value=c.get('date_debut', ''))
            ws.cell(row=row, column=5, value=c.get('date_fin', ''))
            ws.cell(row=row, column=6, value=c.get('transport_type', ''))
            ws.cell(row=row, column=7, value="Oui" if c.get('avec_gasoil', True) else "Non")
            ws.cell(row=row, column=8, value=c.get('statut', ''))
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=chantiers_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    else:
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(["Référence", "Client", "Lieu", "Date Début", "Date Fin", "Type Transport", "Gasoil", "Statut"])
        
        for c in chantiers:
            writer.writerow([
                c.get('reference', ''),
                c.get('client_nom', ''),
                c.get('lieu', ''),
                c.get('date_debut', ''),
                c.get('date_fin', ''),
                c.get('transport_type', ''),
                "Oui" if c.get('avec_gasoil', True) else "Non",
                c.get('statut', '')
            ])
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=chantiers_{datetime.now().strftime('%Y%m%d')}.csv"}
        )
