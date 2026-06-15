from fastapi import FastAPI, APIRouter, HTTPException, Query, Depends
from fastapi.responses import FileResponse, RedirectResponse, Response, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, date, timedelta
from enum import Enum
import io
import base64
import csv

# Auth imports
import bcrypt
from jose import jwt, JWTError

# DocuSign imports
from docusign_esign import ApiClient, EnvelopesApi, EnvelopeDefinition, Document, Signer, SignHere, Tabs, Recipients

# PDF generation
from weasyprint import HTML, CSS

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'TdB-ERP-SecureKey-2026-X9kL3mNp7qRs')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Security
security = HTTPBearer(auto_error=False)

# DocuSign configuration
DOCUSIGN_INTEGRATION_KEY = os.environ.get('DOCUSIGN_INTEGRATION_KEY', 'e54a9490-73fa-4ed9-95ee-0792ccd27423')
DOCUSIGN_SECRET_KEY = os.environ.get('DOCUSIGN_SECRET_KEY', 'b147c93b-24e4-4c7d-ac89-4deadeec97fb')
DOCUSIGN_ACCOUNT_ID = os.environ.get('DOCUSIGN_ACCOUNT_ID', 'bf342295-257f-41bf-8b56-97b8a66f6358')
DOCUSIGN_AUTH_SERVER = os.environ.get('DOCUSIGN_AUTH_SERVER', 'account-d.docusign.com')
DOCUSIGN_BASE_URL = os.environ.get('DOCUSIGN_BASE_URL', 'https://demo.docusign.net/restapi')
# Note: Pour la production, changer AUTH_SERVER -> account.docusign.com et BASE_URL -> https://na4.docusign.net/restapi

# Create the main app without a prefix
app = FastAPI(title="Terre de Beauce ERP")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ============= ENUMS =============
class VehicleStatus(str, Enum):
    DISPONIBLE = "disponible"
    EN_MISSION = "en_mission"
    MAINTENANCE = "maintenance"

class EquipmentType(str, Enum):
    REMORQUE = "remorque"
    CITERNE = "citerne"
    BENNE = "benne"

class ChantierStatus(str, Enum):
    PLANIFIE = "planifie"
    EN_COURS = "en_cours"
    TERMINE = "termine"
    ANNULE = "annule"

class MethodeFacturation(str, Enum):
    HEURE = "heure"
    TONNE = "tonne"
    JOURNEE = "journee"

class FactureStatus(str, Enum):
    BROUILLON = "brouillon"
    EMISE = "emise"
    ENVOYEE = "envoyee"
    SIGNEE = "signee"
    PAYEE = "payee"

class ContratStatus(str, Enum):
    BROUILLON = "brouillon"
    ENVOYE = "envoye"
    SIGNE = "signe"
    ANNULE = "annule"

class TransportType(str, Enum):
    SOLIDE = "solide"
    LIQUIDE = "liquide"

# ============= MODELS =============

# Configuration Entreprise
class EntrepriseConfig(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "config_entreprise"
    raison_sociale: str = "Terre de Beauce"
    adresse: str = "Ferme de Mennessard"
    code_postal: str = "91660"
    ville: str = "Le Mérévillois"
    pays: str = "France"
    siren: str = "953286333"
    siret: str = "95328633300018"
    tva_intracommunautaire: str = "FR57953286333"
    email: str = "r.coisnon@terredebeauce.com"
    telephone: Optional[str] = None
    iban: Optional[str] = None
    bic: Optional[str] = None
    logo_url: Optional[str] = None

# Compte Bancaire
class CompteBancaireBase(BaseModel):
    nom_banque: str
    iban: str
    bic: str
    is_default: bool = False

class CompteBancaireCreate(CompteBancaireBase):
    pass

class CompteBancaire(CompteBancaireBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EntrepriseConfigUpdate(BaseModel):
    raison_sociale: Optional[str] = None
    adresse: Optional[str] = None
    code_postal: Optional[str] = None
    ville: Optional[str] = None
    pays: Optional[str] = None
    siren: Optional[str] = None
    siret: Optional[str] = None
    tva_intracommunautaire: Optional[str] = None
    email: Optional[str] = None
    telephone: Optional[str] = None
    iban: Optional[str] = None
    bic: Optional[str] = None

# Tracteurs
class TracteurBase(BaseModel):
    identifiant: str
    marque: str
    modele: str
    immatriculation: str
    annee: Optional[int] = None
    statut: VehicleStatus = VehicleStatus.DISPONIBLE
    notes: Optional[str] = None

class TracteurCreate(TracteurBase):
    pass

class Tracteur(TracteurBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Equipements (Remorques, Citernes, Bennes)
class EquipementBase(BaseModel):
    numero: str
    type: EquipmentType
    capacite: Optional[str] = None
    marque: Optional[str] = None
    immatriculation: Optional[str] = None
    statut: VehicleStatus = VehicleStatus.DISPONIBLE
    notes: Optional[str] = None

class EquipementCreate(EquipementBase):
    pass

class Equipement(EquipementBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Chauffeurs
class ChauffeurBase(BaseModel):
    nom: str
    prenom: str
    telephone: str
    email: Optional[str] = None
    permis: str
    date_embauche: Optional[str] = None
    disponible: bool = True
    code_acces: Optional[str] = None  # Code pour accès chauffeur
    notes: Optional[str] = None

class ChauffeurCreate(ChauffeurBase):
    pass

class Chauffeur(ChauffeurBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Tarification Client
class TarifClient(BaseModel):
    methode: MethodeFacturation
    prix_unitaire: float  # €/h, €/tonne ou €/jour
    description: Optional[str] = None

class ClientBase(BaseModel):
    raison_sociale: str
    siren: Optional[str] = None
    siret: Optional[str] = None
    tva_intracommunautaire: Optional[str] = None
    adresse: str
    code_postal: str
    ville: str
    pays: str = "France"
    telephone: Optional[str] = None
    email: Optional[str] = None
    contact_nom: Optional[str] = None
    contact_telephone: Optional[str] = None
    # Tarification
    tarifs: List[TarifClient] = []
    notes: Optional[str] = None

class ClientCreate(ClientBase):
    pass

class Client(ClientBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Affectation (pour un chantier) - Simplifiée
class Affectation(BaseModel):
    tracteur_id: str
    tracteur_identifiant: Optional[str] = None
    equipement_id: str
    equipement_numero: Optional[str] = None
    equipement_type: Optional[str] = None
    chauffeur_id: str
    chauffeur_nom: Optional[str] = None

# Tarification Chantier
class TarifChantier(BaseModel):
    methode: MethodeFacturation
    prix_unitaire: float
    description: Optional[str] = None

# Chantiers
class ChantierBase(BaseModel):
    reference: str = ""  # Généré automatiquement côté serveur
    client_id: str
    client_nom: Optional[str] = None
    lieu: str
    lieu_chargement: Optional[str] = None
    lieu_dechargement: Optional[str] = None
    description: Optional[str] = None
    date_debut: str
    date_fin: Optional[str] = None
    statut: ChantierStatus = ChantierStatus.PLANIFIE
    affectations: List[Affectation] = []
    # Tarification du chantier
    tarifs: List[TarifChantier] = []
    transport_type: TransportType = TransportType.SOLIDE
    avec_gasoil: bool = True  # True = gasoil fourni, False = sans gasoil
    facturation_kilometrique: bool = False  # True = facturation au km
    notes: Optional[str] = None

class ChantierCreate(ChantierBase):
    pass

class Chantier(ChantierBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Tour (rotation avec volume et distance)
class Tour(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    heure: Optional[str] = None  # Heure du tour
    volume: float = 0  # Tonnes ou m³
    distance_km: float = 0  # Distance chargement -> déchargement
    lieu_chargement: Optional[str] = None
    lieu_dechargement: Optional[str] = None
    commentaire: Optional[str] = None

# Pointage Chauffeur (Saisie heures et volumes)
class PointageBase(BaseModel):
    chantier_id: str
    chauffeur_id: str
    date: str  # YYYY-MM-DD
    heures_travaillees: float = 0
    tours: List[Tour] = []  # Liste des tours de la journée
    commentaire: Optional[str] = None

class PointageCreate(BaseModel):
    chantier_id: str
    chauffeur_id: str
    date: str
    heures_travaillees: float = 0
    tours: List[Tour] = []
    commentaire: Optional[str] = None

class Pointage(PointageBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    chantier_reference: Optional[str] = None
    chauffeur_nom: Optional[str] = None
    client_nom: Optional[str] = None
    transport_type: Optional[str] = None
    avec_gasoil: Optional[bool] = None
    # Totaux calculés
    total_volume: float = 0
    total_distance: float = 0
    nombre_tours: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Barème kilométrique
class TrancheBareme(BaseModel):
    km_min: float
    km_max: Optional[float] = None  # None = illimité
    prix_tonne_km: float  # €/tonne.km ou €/m³.km

class Bareme(BaseModel):
    tranches: List[TrancheBareme] = []

class BaremesConfig(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "config_baremes"
    # Barèmes solide
    solide_avec_gasoil: Bareme = Field(default_factory=Bareme)
    solide_sans_gasoil: Bareme = Field(default_factory=Bareme)
    # Barèmes liquide
    liquide_avec_gasoil: Bareme = Field(default_factory=Bareme)
    liquide_sans_gasoil: Bareme = Field(default_factory=Bareme)
    # Taux horaire minimum
    taux_horaire_minimum: float = 0
    # Unités
    unite_solide: str = "tonnes"
    unite_liquide: str = "m³"

class BaremesConfigUpdate(BaseModel):
    solide_avec_gasoil: Optional[Bareme] = None
    solide_sans_gasoil: Optional[Bareme] = None
    liquide_avec_gasoil: Optional[Bareme] = None
    liquide_sans_gasoil: Optional[Bareme] = None
    taux_horaire_minimum: Optional[float] = None
    unite_solide: Optional[str] = None
    unite_liquide: Optional[str] = None

# Ligne de facture
class LigneFacture(BaseModel):
    description: str
    quantite: float
    unite: str  # "heures", "tonnes", "jours"
    prix_unitaire: float
    montant_ht: float

# Photo attachée à un pointage
class PhotoPointage(BaseModel):
    id: str
    url: str
    type: str = "pointage"  # "pointage" ou "note_frais"
    description: Optional[str] = None
    date_creation: str

# Note de frais
class NoteFrais(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    chauffeur_id: str
    chauffeur_nom: Optional[str] = None
    pointage_id: Optional[str] = None
    chantier_id: Optional[str] = None
    chantier_reference: Optional[str] = None
    date: str
    montant: float
    type_frais: str  # "carburant", "peage", "repas", "hebergement", "autre"
    description: Optional[str] = None
    photo_url: Optional[str] = None
    statut: str = "en_attente"  # "en_attente", "valide", "refuse", "rembourse"
    date_creation: str

class NoteFraisCreate(BaseModel):
    chauffeur_id: str
    pointage_id: Optional[str] = None
    chantier_id: Optional[str] = None
    date: str
    montant: float
    type_frais: str
    description: Optional[str] = None
    photo_base64: Optional[str] = None

# Facture
class FactureBase(BaseModel):
    chantier_id: str
    client_id: str
    numero: str
    date_emission: str
    date_echeance: str
    lignes: List[LigneFacture] = []
    montant_ht: float = 0
    taux_tva: float = 20.0
    montant_tva: float = 0
    montant_ttc: float = 0
    statut: FactureStatus = FactureStatus.BROUILLON
    notes: Optional[str] = None
    # Infos pour PDF
    client_raison_sociale: Optional[str] = None
    client_adresse: Optional[str] = None
    client_siret: Optional[str] = None
    client_tva: Optional[str] = None
    client_email: Optional[str] = None
    chantier_reference: Optional[str] = None
    chantier_lieu: Optional[str] = None
    # Lien avec contrat CCPA
    contrat_id: Optional[str] = None
    contrat_numero: Optional[str] = None
    # Compte bancaire pour le paiement
    compte_bancaire_id: Optional[str] = None
    compte_bancaire_nom: Optional[str] = None
    compte_bancaire_iban: Optional[str] = None
    compte_bancaire_bic: Optional[str] = None

class FactureCreate(BaseModel):
    chantier_id: str
    date_echeance: str
    notes: Optional[str] = None
    compte_bancaire_id: Optional[str] = None

class Facture(FactureBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Contrat CCPA (Contrat Cadre de Prestations Agricoles)
class ContratCCPA(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    numero_contrat: str  # Numéro unique du contrat
    chantier_id: str  # Lien avec le chantier
    client_id: str
    
    # Informations client (modifiables - zones jaunes)
    client_nom: str = ""
    client_interlocuteur: str = ""
    client_adresse: str = ""
    client_email: str = ""
    client_telephone: str = ""
    
    # Tarification (modifiable - zone jaune)
    prix_unitaire: float = 0
    unite_facturation: str = ""  # "heure", "tonne", "m³", "jour", etc.
    
    # Option gasoil (héritée du chantier)
    gasoil_fourni: bool = True
    transport_type: TransportType = TransportType.SOLIDE
    
    # Dates
    date_creation: str = ""
    date_signature: Optional[str] = None
    
    # Statut
    statut: ContratStatus = ContratStatus.BROUILLON
    
    # DocuSign
    docusign_envelope_id: Optional[str] = None
    
    # Métadonnées
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ContratCCPACreate(BaseModel):
    chantier_id: str
    # Les autres champs sont pré-remplis depuis le chantier/client

class ContratCCPAUpdate(BaseModel):
    client_nom: Optional[str] = None
    client_interlocuteur: Optional[str] = None
    client_adresse: Optional[str] = None
    client_email: Optional[str] = None
    client_telephone: Optional[str] = None
    prix_unitaire: Optional[float] = None
    unite_facturation: Optional[str] = None
    date_signature: Optional[str] = None
    statut: Optional[ContratStatus] = None

# Ancien modèle de contrat (gardé pour compatibilité)
class ContratBase(BaseModel):
    client_id: str
    type_transport: TransportType
    reference: str
    date_creation: str
    conditions: Optional[str] = None
    statut: ContratStatus = ContratStatus.BROUILLON
    # Infos client
    client_raison_sociale: Optional[str] = None
    client_adresse: Optional[str] = None
    client_siret: Optional[str] = None
    client_email: Optional[str] = None
    # Signature
    docusign_envelope_id: Optional[str] = None
    date_signature: Optional[str] = None

class ContratCreate(BaseModel):
    client_id: str
    type_transport: TransportType
    conditions: Optional[str] = None

class Contrat(ContratBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Dashboard Stats
class DashboardStats(BaseModel):
    total_tracteurs: int
    tracteurs_disponibles: int
    total_equipements: int
    equipements_disponibles: int
    total_chauffeurs: int
    chauffeurs_disponibles: int
    total_clients: int
    chantiers_en_cours: int
    chantiers_planifies: int
    factures_en_attente: int = 0
    ca_mois: float = 0

# Auth Chauffeur
class ChauffeurLogin(BaseModel):
    code_acces: str

class ChauffeurSession(BaseModel):
    chauffeur_id: str
    chauffeur_nom: str
    token: str

# ============= ADMIN AUTH MODELS =============
class AdminBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    email: str
    nom: str
    prenom: str
    role: str = "admin"

class AdminCreate(AdminBase):
    password: str

class Admin(AdminBase):
    id: str
    date_creation: str
    is_active: bool = True

class AdminLogin(BaseModel):
    email: str
    password: str

class AdminSession(BaseModel):
    admin_id: str
    email: str
    nom: str
    prenom: str
    token: str

# ============= AUTH HELPER FUNCTIONS =============
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_jwt_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_jwt_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except JWTError:
        return None

async def get_current_admin(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not credentials:
        raise HTTPException(status_code=401, detail="Non authentifié")
    
    payload = decode_jwt_token(credentials.credentials)
    if not payload or payload.get("type") != "admin":
        raise HTTPException(status_code=401, detail="Token invalide")
    
    admin = await db.admins.find_one({"id": payload.get("admin_id")}, {"_id": 0})
    if not admin or not admin.get("is_active"):
        raise HTTPException(status_code=401, detail="Administrateur non trouvé ou désactivé")
    
    return admin

# ============= HELPER FUNCTIONS =============
def serialize_doc(doc: dict) -> dict:
    if '_id' in doc:
        del doc['_id']
    if 'created_at' in doc and isinstance(doc['created_at'], datetime):
        doc['created_at'] = doc['created_at'].isoformat()
    if 'updated_at' in doc and isinstance(doc['updated_at'], datetime):
        doc['updated_at'] = doc['updated_at'].isoformat()
    return doc

async def generate_numero_chantier() -> str:
    """
    Génère un numéro de chantier séquentiel par année : CH-0003-2026
    Utilise un compteur atomique en base (collection 'counters').
    Le compteur 2026 démarre à 3 (deux factures déjà émises).
    """
    annee = datetime.now().year
    counter_id = f"chantier_{annee}"
    # Initialise à 2 si n'existe pas (incrément donnera 3 au premier appel)
    result = await db.counters.find_one_and_update(
        {"_id": counter_id},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    seq = result["seq"]
    # Si c'est le premier document 2026, on s'assure de partir à 3
    if annee == 2026 and seq < 3:
        await db.counters.update_one(
            {"_id": counter_id},
            {"$set": {"seq": 3}}
        )
        seq = 3
    return f"CH-{seq:04d}-{annee}"

def generate_numero_contrat_from_chantier(reference_chantier: str) -> str:
    """CT-0003-2026 depuis CH-0003-2026"""
    return reference_chantier.replace("CH-", "CT-", 1)

def generate_numero_releve_from_chantier(reference_chantier: str) -> str:
    """RH-0003-2026 depuis CH-0003-2026"""
    return reference_chantier.replace("CH-", "RH-", 1)

def generate_numero_facture_from_chantier(reference_chantier: str) -> str:
    """FC-0003-2026 depuis CH-0003-2026"""
    return reference_chantier.replace("CH-", "FC-", 1)

# Conservé pour compatibilité avec l'ancien modèle de contrat (non-CCPA)
def generate_numero_facture():
    now = datetime.now()
    return f"FAC-{now.year}-{now.month:02d}-{str(uuid.uuid4())[:8].upper()}"

def generate_reference_contrat(type_transport: str):
    now = datetime.now()
    prefix = "CTR-S" if type_transport == "solide" else "CTR-L"
    return f"{prefix}-{now.year}-{str(uuid.uuid4())[:6].upper()}"

def generate_numero_contrat_ccpa():
    """Conservé pour compatibilité — préférer generate_numero_contrat_from_chantier"""
    now = datetime.now()
    return f"CCPA-{now.year}-{now.month:02d}-{str(uuid.uuid4())[:6].upper()}"

# ============= ADMIN AUTH ROUTES =============
@api_router.post("/admin/login")
async def admin_login(login: AdminLogin):
    """Connexion administrateur"""
    admin = await db.admins.find_one({"email": login.email.lower()}, {"_id": 0})
    if not admin:
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    if not verify_password(login.password, admin.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    if not admin.get("is_active", True):
        raise HTTPException(status_code=401, detail="Compte désactivé")
    
    token = create_jwt_token({
        "admin_id": admin["id"],
        "email": admin["email"],
        "type": "admin"
    })
    
    return AdminSession(
        admin_id=admin["id"],
        email=admin["email"],
        nom=admin["nom"],
        prenom=admin["prenom"],
        token=token
    )

@api_router.get("/admin/me")
async def get_current_admin_info(admin: dict = Depends(get_current_admin)):
    """Récupère les infos de l'admin connecté"""
    return {
        "id": admin["id"],
        "email": admin["email"],
        "nom": admin["nom"],
        "prenom": admin["prenom"],
        "role": admin.get("role", "admin")
    }

@api_router.get("/admin/list")
async def list_admins(admin: dict = Depends(get_current_admin)):
    """Liste tous les administrateurs"""
    admins = await db.admins.find({}, {"_id": 0, "password_hash": 0}).to_list(100)
    return admins

@api_router.post("/admin/create")
async def create_admin(new_admin: AdminCreate, admin: dict = Depends(get_current_admin)):
    """Créer un nouvel administrateur"""
    # Vérifier si l'email existe déjà
    existing = await db.admins.find_one({"email": new_admin.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Un administrateur avec cet email existe déjà")
    
    admin_dict = {
        "id": str(uuid.uuid4())[:8].upper(),
        "email": new_admin.email.lower(),
        "nom": new_admin.nom,
        "prenom": new_admin.prenom,
        "role": new_admin.role,
        "password_hash": hash_password(new_admin.password),
        "is_active": True,
        "date_creation": datetime.now(timezone.utc).isoformat()
    }
    
    await db.admins.insert_one(admin_dict)
    
    # Retourner sans le hash et sans _id (ajouté par MongoDB)
    del admin_dict["password_hash"]
    if "_id" in admin_dict:
        del admin_dict["_id"]
    return admin_dict

@api_router.delete("/admin/{admin_id}")
async def delete_admin(admin_id: str, admin: dict = Depends(get_current_admin)):
    """Supprimer un administrateur"""
    if admin_id == admin["id"]:
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas supprimer votre propre compte")
    
    result = await db.admins.delete_one({"id": admin_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Administrateur non trouvé")
    
    return {"message": "Administrateur supprimé"}

@api_router.put("/admin/{admin_id}/toggle-active")
async def toggle_admin_active(admin_id: str, admin: dict = Depends(get_current_admin)):
    """Activer/désactiver un administrateur"""
    if admin_id == admin["id"]:
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas désactiver votre propre compte")
    
    target_admin = await db.admins.find_one({"id": admin_id}, {"_id": 0})
    if not target_admin:
        raise HTTPException(status_code=404, detail="Administrateur non trouvé")
    
    new_status = not target_admin.get("is_active", True)
    await db.admins.update_one({"id": admin_id}, {"$set": {"is_active": new_status}})
    
    return {"message": f"Administrateur {'activé' if new_status else 'désactivé'}", "is_active": new_status}

@api_router.post("/admin/init")
async def init_admin():
    """Initialise le premier administrateur si aucun n'existe"""
    count = await db.admins.count_documents({})
    if count > 0:
        raise HTTPException(status_code=400, detail="Des administrateurs existent déjà")
    
    # Créer l'admin par défaut
    admin_dict = {
        "id": str(uuid.uuid4())[:8].upper(),
        "email": "r.coisnon@terredebeauce.com",
        "nom": "Coisnon",
        "prenom": "Romain",
        "role": "admin",
        "password_hash": hash_password("Mennessard03"),
        "is_active": True,
        "date_creation": datetime.now(timezone.utc).isoformat()
    }
    
    await db.admins.insert_one(admin_dict)
    
    return {"message": "Administrateur initial créé", "email": admin_dict["email"]}

@api_router.get("/admin/check")
async def check_admin_exists():
    """Vérifie si des administrateurs existent"""
    count = await db.admins.count_documents({})
    return {"has_admins": count > 0}

# ============= ENTREPRISE CONFIG ROUTES =============
@api_router.get("/config/entreprise", response_model=EntrepriseConfig)
async def get_entreprise_config():
    config = await db.config.find_one({"id": "config_entreprise"}, {"_id": 0})
    if not config:
        # Créer config par défaut
        default_config = EntrepriseConfig()
        await db.config.insert_one(default_config.model_dump())
        return default_config
    return config

@api_router.put("/config/entreprise", response_model=EntrepriseConfig)
async def update_entreprise_config(input: EntrepriseConfigUpdate):
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    await db.config.update_one(
        {"id": "config_entreprise"},
        {"$set": update_data},
        upsert=True
    )
    config = await db.config.find_one({"id": "config_entreprise"}, {"_id": 0})
    return config

# ============= BAREMES KILOMÉTRIQUES ROUTES =============
def generate_default_baremes() -> BaremesConfig:
    """Génère les barèmes par défaut avec tranches de 2.5km jusqu'à 50km"""
    tranches = []
    for i in range(20):  # 20 tranches de 2.5km = 50km
        km_min = i * 2.5
        km_max = (i + 1) * 2.5
        tranches.append(TrancheBareme(km_min=km_min, km_max=km_max, prix_tonne_km=0))
    
    bareme = Bareme(tranches=tranches)
    return BaremesConfig(
        solide_avec_gasoil=bareme,
        solide_sans_gasoil=bareme,
        liquide_avec_gasoil=bareme,
        liquide_sans_gasoil=bareme,
        taux_horaire_minimum=0
    )

@api_router.get("/config/baremes")
async def get_baremes_config():
    """Récupère la configuration des barèmes kilométriques"""
    config = await db.config.find_one({"id": "config_baremes"}, {"_id": 0})
    if not config:
        # Créer les barèmes par défaut
        default_baremes = generate_default_baremes()
        doc = default_baremes.model_dump()
        await db.config.insert_one(doc)
        # Récupérer le doc sans _id
        config = await db.config.find_one({"id": "config_baremes"}, {"_id": 0})
        return config
    return config

@api_router.put("/config/baremes")
async def update_baremes_config(input: BaremesConfigUpdate):
    """Met à jour la configuration des barèmes kilométriques"""
    # Récupérer les barèmes existants ou créer par défaut
    existing = await db.config.find_one({"id": "config_baremes"}, {"_id": 0})
    if not existing:
        default_baremes = generate_default_baremes()
        existing = default_baremes.model_dump()
        await db.config.insert_one(existing)
    
    # Mettre à jour uniquement les champs fournis
    update_data = {}
    input_data = input.model_dump(exclude_unset=True)
    
    for key, value in input_data.items():
        if value is not None:
            if isinstance(value, dict) and 'tranches' in value:
                update_data[key] = value
            else:
                update_data[key] = value
    
    if update_data:
        await db.config.update_one(
            {"id": "config_baremes"},
            {"$set": update_data},
            upsert=True
        )
    
    updated = await db.config.find_one({"id": "config_baremes"}, {"_id": 0})
    return updated

@api_router.put("/config/baremes/{bareme_type}")
async def update_single_bareme(bareme_type: str, bareme: Bareme):
    """Met à jour un seul barème (solide_avec_gasoil, solide_sans_gasoil, etc.)"""
    valid_types = ["solide_avec_gasoil", "solide_sans_gasoil", "liquide_avec_gasoil", "liquide_sans_gasoil"]
    if bareme_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Type de barème invalide. Valeurs acceptées: {valid_types}")
    
    # S'assurer que les barèmes existent
    existing = await db.config.find_one({"id": "config_baremes"}, {"_id": 0})
    if not existing:
        default_baremes = generate_default_baremes()
        await db.config.insert_one(default_baremes.model_dump())
    
    await db.config.update_one(
        {"id": "config_baremes"},
        {"$set": {bareme_type: bareme.model_dump()}}
    )
    
    updated = await db.config.find_one({"id": "config_baremes"}, {"_id": 0})
    return updated

@api_router.put("/config/baremes/taux-horaire-minimum")
async def update_taux_horaire_minimum(taux: float):
    """Met à jour le taux horaire minimum"""
    existing = await db.config.find_one({"id": "config_baremes"}, {"_id": 0})
    if not existing:
        default_baremes = generate_default_baremes()
        await db.config.insert_one(default_baremes.model_dump())
    
    await db.config.update_one(
        {"id": "config_baremes"},
        {"$set": {"taux_horaire_minimum": taux}}
    )
    
    updated = await db.config.find_one({"id": "config_baremes"}, {"_id": 0})
    return updated

# ============= COMPTES BANCAIRES ROUTES =============
@api_router.get("/comptes-bancaires", response_model=List[CompteBancaire])
async def get_comptes_bancaires():
    """Récupère tous les comptes bancaires"""
    comptes = await db.comptes_bancaires.find({}, {"_id": 0}).to_list(100)
    return comptes

@api_router.get("/comptes-bancaires/{compte_id}", response_model=CompteBancaire)
async def get_compte_bancaire(compte_id: str):
    """Récupère un compte bancaire par son ID"""
    compte = await db.comptes_bancaires.find_one({"id": compte_id}, {"_id": 0})
    if not compte:
        raise HTTPException(status_code=404, detail="Compte bancaire non trouvé")
    return compte

@api_router.post("/comptes-bancaires", response_model=CompteBancaire)
async def create_compte_bancaire(input: CompteBancaireCreate):
    """Crée un nouveau compte bancaire"""
    compte = CompteBancaire(**input.model_dump())
    doc = compte.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    # Si c'est le premier compte ou si is_default est True, gérer le défaut
    if input.is_default:
        # Retirer le défaut des autres comptes
        await db.comptes_bancaires.update_many({}, {"$set": {"is_default": False}})
    else:
        # Si c'est le premier compte, le mettre par défaut
        existing_count = await db.comptes_bancaires.count_documents({})
        if existing_count == 0:
            doc['is_default'] = True
    
    await db.comptes_bancaires.insert_one(doc)
    return compte

@api_router.put("/comptes-bancaires/{compte_id}", response_model=CompteBancaire)
async def update_compte_bancaire(compte_id: str, input: CompteBancaireCreate):
    """Met à jour un compte bancaire"""
    existing = await db.comptes_bancaires.find_one({"id": compte_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Compte bancaire non trouvé")
    
    update_data = input.model_dump()
    
    # Gérer le compte par défaut
    if input.is_default:
        await db.comptes_bancaires.update_many(
            {"id": {"$ne": compte_id}},
            {"$set": {"is_default": False}}
        )
    
    await db.comptes_bancaires.update_one({"id": compte_id}, {"$set": update_data})
    updated = await db.comptes_bancaires.find_one({"id": compte_id}, {"_id": 0})
    return updated

@api_router.delete("/comptes-bancaires/{compte_id}")
async def delete_compte_bancaire(compte_id: str):
    """Supprime un compte bancaire"""
    existing = await db.comptes_bancaires.find_one({"id": compte_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Compte bancaire non trouvé")
    
    result = await db.comptes_bancaires.delete_one({"id": compte_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Compte bancaire non trouvé")
    
    # Si le compte supprimé était par défaut, mettre un autre compte par défaut
    if existing.get('is_default'):
        first_compte = await db.comptes_bancaires.find_one({}, {"_id": 0})
        if first_compte:
            await db.comptes_bancaires.update_one(
                {"id": first_compte['id']},
                {"$set": {"is_default": True}}
            )
    
    return {"message": "Compte bancaire supprimé"}

# ============= TRACTEURS ROUTES =============
@api_router.get("/tracteurs", response_model=List[Tracteur])
async def get_tracteurs():
    tracteurs = await db.tracteurs.find({}, {"_id": 0}).to_list(1000)
    return tracteurs

@api_router.get("/tracteurs/{tracteur_id}", response_model=Tracteur)
async def get_tracteur(tracteur_id: str):
    tracteur = await db.tracteurs.find_one({"id": tracteur_id}, {"_id": 0})
    if not tracteur:
        raise HTTPException(status_code=404, detail="Tracteur non trouvé")
    return tracteur

@api_router.post("/tracteurs", response_model=Tracteur)
async def create_tracteur(input: TracteurCreate):
    tracteur = Tracteur(**input.model_dump())
    doc = tracteur.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.tracteurs.insert_one(doc)
    return tracteur

@api_router.put("/tracteurs/{tracteur_id}", response_model=Tracteur)
async def update_tracteur(tracteur_id: str, input: TracteurCreate):
    existing = await db.tracteurs.find_one({"id": tracteur_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Tracteur non trouvé")
    update_data = input.model_dump()
    await db.tracteurs.update_one({"id": tracteur_id}, {"$set": update_data})
    updated = await db.tracteurs.find_one({"id": tracteur_id}, {"_id": 0})
    return updated

@api_router.delete("/tracteurs/{tracteur_id}")
async def delete_tracteur(tracteur_id: str):
    result = await db.tracteurs.delete_one({"id": tracteur_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tracteur non trouvé")
    return {"message": "Tracteur supprimé"}

# ============= EQUIPEMENTS ROUTES =============
@api_router.get("/equipements", response_model=List[Equipement])
async def get_equipements(type: Optional[EquipmentType] = None):
    query = {}
    if type:
        query["type"] = type
    equipements = await db.equipements.find(query, {"_id": 0}).to_list(1000)
    return equipements

@api_router.get("/equipements/{equipement_id}", response_model=Equipement)
async def get_equipement(equipement_id: str):
    equipement = await db.equipements.find_one({"id": equipement_id}, {"_id": 0})
    if not equipement:
        raise HTTPException(status_code=404, detail="Équipement non trouvé")
    return equipement

@api_router.post("/equipements", response_model=Equipement)
async def create_equipement(input: EquipementCreate):
    equipement = Equipement(**input.model_dump())
    doc = equipement.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.equipements.insert_one(doc)
    return equipement

@api_router.put("/equipements/{equipement_id}", response_model=Equipement)
async def update_equipement(equipement_id: str, input: EquipementCreate):
    existing = await db.equipements.find_one({"id": equipement_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Équipement non trouvé")
    update_data = input.model_dump()
    await db.equipements.update_one({"id": equipement_id}, {"$set": update_data})
    updated = await db.equipements.find_one({"id": equipement_id}, {"_id": 0})
    return updated

@api_router.delete("/equipements/{equipement_id}")
async def delete_equipement(equipement_id: str):
    result = await db.equipements.delete_one({"id": equipement_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Équipement non trouvé")
    return {"message": "Équipement supprimé"}

# ============= CHAUFFEURS ROUTES =============
@api_router.get("/chauffeurs", response_model=List[Chauffeur])
async def get_chauffeurs():
    chauffeurs = await db.chauffeurs.find({}, {"_id": 0}).to_list(1000)
    return chauffeurs

@api_router.get("/chauffeurs/{chauffeur_id}", response_model=Chauffeur)
async def get_chauffeur(chauffeur_id: str):
    chauffeur = await db.chauffeurs.find_one({"id": chauffeur_id}, {"_id": 0})
    if not chauffeur:
        raise HTTPException(status_code=404, detail="Chauffeur non trouvé")
    return chauffeur

@api_router.post("/chauffeurs", response_model=Chauffeur)
async def create_chauffeur(input: ChauffeurCreate):
    chauffeur_data = input.model_dump()
    # Générer un code d'accès automatique si non fourni
    if not chauffeur_data.get('code_acces'):
        chauffeur_data['code_acces'] = str(uuid.uuid4())[:6].upper()
    chauffeur = Chauffeur(**chauffeur_data)
    doc = chauffeur.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.chauffeurs.insert_one(doc)
    return chauffeur

@api_router.put("/chauffeurs/{chauffeur_id}", response_model=Chauffeur)
async def update_chauffeur(chauffeur_id: str, input: ChauffeurCreate):
    existing = await db.chauffeurs.find_one({"id": chauffeur_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Chauffeur non trouvé")
    update_data = input.model_dump()
    await db.chauffeurs.update_one({"id": chauffeur_id}, {"$set": update_data})
    updated = await db.chauffeurs.find_one({"id": chauffeur_id}, {"_id": 0})
    return updated

@api_router.delete("/chauffeurs/{chauffeur_id}")
async def delete_chauffeur(chauffeur_id: str):
    result = await db.chauffeurs.delete_one({"id": chauffeur_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Chauffeur non trouvé")
    return {"message": "Chauffeur supprimé"}

# ============= AUTH CHAUFFEUR =============
@api_router.post("/chauffeur/login", response_model=ChauffeurSession)
async def chauffeur_login(input: ChauffeurLogin):
    chauffeur = await db.chauffeurs.find_one({"code_acces": input.code_acces}, {"_id": 0})
    if not chauffeur:
        raise HTTPException(status_code=401, detail="Code d'accès invalide")
    
    # Générer un token simple
    token = str(uuid.uuid4())
    
    return ChauffeurSession(
        chauffeur_id=chauffeur['id'],
        chauffeur_nom=f"{chauffeur['prenom']} {chauffeur['nom']}",
        token=token
    )

# ============= CLIENTS ROUTES =============
@api_router.get("/clients", response_model=List[Client])
async def get_clients():
    clients = await db.clients.find({}, {"_id": 0}).to_list(1000)
    return clients

@api_router.get("/clients/{client_id}", response_model=Client)
async def get_client(client_id: str):
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    return client

@api_router.post("/clients", response_model=Client)
async def create_client(input: ClientCreate):
    client = Client(**input.model_dump())
    doc = client.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.clients.insert_one(doc)
    return client

@api_router.put("/clients/{client_id}", response_model=Client)
async def update_client(client_id: str, input: ClientCreate):
    existing = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    update_data = input.model_dump()
    await db.clients.update_one({"id": client_id}, {"$set": update_data})
    updated = await db.clients.find_one({"id": client_id}, {"_id": 0})
    return updated

@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str):
    result = await db.clients.delete_one({"id": client_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    return {"message": "Client supprimé"}

# ============= CHANTIERS ROUTES =============
@api_router.get("/chantiers", response_model=List[Chantier])
async def get_chantiers(statut: Optional[ChantierStatus] = None):
    query = {}
    if statut:
        query["statut"] = statut
    chantiers = await db.chantiers.find(query, {"_id": 0}).to_list(1000)
    return chantiers

@api_router.get("/chantiers/{chantier_id}", response_model=Chantier)
async def get_chantier(chantier_id: str):
    chantier = await db.chantiers.find_one({"id": chantier_id}, {"_id": 0})
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier non trouvé")
    return chantier

@api_router.post("/chantiers", response_model=Chantier)
async def create_chantier(input: ChantierCreate):
    # Génération automatique de la référence séquentielle CH-XXXX-YYYY
    reference = await generate_numero_chantier()

    # Enrich affectations with names
    enriched_affectations = []
    for aff in input.affectations:
        aff_dict = aff.model_dump()
        tracteur = await db.tracteurs.find_one({"id": aff.tracteur_id}, {"_id": 0})
        if tracteur:
            aff_dict['tracteur_identifiant'] = tracteur.get('identifiant')
        equipement = await db.equipements.find_one({"id": aff.equipement_id}, {"_id": 0})
        if equipement:
            aff_dict['equipement_numero'] = equipement.get('numero')
            aff_dict['equipement_type'] = equipement.get('type')
        chauffeur = await db.chauffeurs.find_one({"id": aff.chauffeur_id}, {"_id": 0})
        if chauffeur:
            aff_dict['chauffeur_nom'] = f"{chauffeur.get('prenom', '')} {chauffeur.get('nom', '')}".strip()
        enriched_affectations.append(aff_dict)
    
    # Get client name
    client = await db.clients.find_one({"id": input.client_id}, {"_id": 0})
    client_nom = client.get('raison_sociale') if client else None
    
    # Si pas de tarifs définis, récupérer ceux du client
    tarifs = input.tarifs
    if not tarifs and client and client.get('tarifs'):
        tarifs = [TarifChantier(**t) for t in client.get('tarifs', [])]
    
    chantier_data = input.model_dump()
    chantier_data['reference'] = reference  # Écrase la référence saisie par l'auto-générée
    chantier_data['affectations'] = enriched_affectations
    chantier_data['client_nom'] = client_nom
    chantier_data['tarifs'] = [t.model_dump() if hasattr(t, 'model_dump') else t for t in tarifs]
    
    chantier = Chantier(**chantier_data)
    doc = chantier.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.chantiers.insert_one(doc)
    return chantier

@api_router.put("/chantiers/{chantier_id}", response_model=Chantier)
async def update_chantier(chantier_id: str, input: ChantierCreate):
    existing = await db.chantiers.find_one({"id": chantier_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Chantier non trouvé")
    
    enriched_affectations = []
    for aff in input.affectations:
        aff_dict = aff.model_dump()
        tracteur = await db.tracteurs.find_one({"id": aff.tracteur_id}, {"_id": 0})
        if tracteur:
            aff_dict['tracteur_identifiant'] = tracteur.get('identifiant')
        equipement = await db.equipements.find_one({"id": aff.equipement_id}, {"_id": 0})
        if equipement:
            aff_dict['equipement_numero'] = equipement.get('numero')
            aff_dict['equipement_type'] = equipement.get('type')
        chauffeur = await db.chauffeurs.find_one({"id": aff.chauffeur_id}, {"_id": 0})
        if chauffeur:
            aff_dict['chauffeur_nom'] = f"{chauffeur.get('prenom', '')} {chauffeur.get('nom', '')}".strip()
        enriched_affectations.append(aff_dict)
    
    client = await db.clients.find_one({"id": input.client_id}, {"_id": 0})
    client_nom = client.get('raison_sociale') if client else None
    
    update_data = input.model_dump()
    update_data['affectations'] = enriched_affectations
    update_data['client_nom'] = client_nom
    
    await db.chantiers.update_one({"id": chantier_id}, {"$set": update_data})
    updated = await db.chantiers.find_one({"id": chantier_id}, {"_id": 0})
    return updated

@api_router.delete("/chantiers/{chantier_id}")
async def delete_chantier(chantier_id: str):
    result = await db.chantiers.delete_one({"id": chantier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Chantier non trouvé")
    return {"message": "Chantier supprimé"}

# Chantiers pour un chauffeur
@api_router.get("/chauffeur/{chauffeur_id}/chantiers")
async def get_chauffeur_chantiers(chauffeur_id: str):
    # Récupérer les chantiers où le chauffeur est affecté et qui sont en cours ou planifiés
    chantiers = await db.chantiers.find({
        "affectations.chauffeur_id": chauffeur_id,
        "statut": {"$in": ["planifie", "en_cours"]}
    }, {"_id": 0}).to_list(100)
    return chantiers

# ============= POINTAGES ROUTES =============
@api_router.get("/pointages", response_model=List[Pointage])
async def get_pointages(chantier_id: Optional[str] = None, chauffeur_id: Optional[str] = None, date: Optional[str] = None):
    query = {}
    if chantier_id:
        query["chantier_id"] = chantier_id
    if chauffeur_id:
        query["chauffeur_id"] = chauffeur_id
    if date:
        query["date"] = date
    pointages = await db.pointages.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return pointages

@api_router.get("/pointages/{pointage_id}", response_model=Pointage)
async def get_pointage(pointage_id: str):
    pointage = await db.pointages.find_one({"id": pointage_id}, {"_id": 0})
    if not pointage:
        raise HTTPException(status_code=404, detail="Pointage non trouvé")
    return pointage

@api_router.post("/pointages", response_model=Pointage)
async def create_pointage(input: PointageCreate):
    # Vérifier que le chauffeur est bien affecté au chantier
    chantier = await db.chantiers.find_one({"id": input.chantier_id}, {"_id": 0})
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier non trouvé")
    
    chauffeur = await db.chauffeurs.find_one({"id": input.chauffeur_id}, {"_id": 0})
    if not chauffeur:
        raise HTTPException(status_code=404, detail="Chauffeur non trouvé")
    
    # Calculer les totaux depuis les tours
    tours = input.tours or []
    total_volume = sum(t.volume for t in tours)
    total_distance = sum(t.distance_km for t in tours)
    nombre_tours = len(tours)
    
    # Vérifier si un pointage existe déjà pour ce jour
    existing = await db.pointages.find_one({
        "chantier_id": input.chantier_id,
        "chauffeur_id": input.chauffeur_id,
        "date": input.date
    }, {"_id": 0})
    
    if existing:
        # Mettre à jour le pointage existant
        update_data = input.model_dump()
        update_data['tours'] = [t.model_dump() if hasattr(t, 'model_dump') else t for t in tours]
        update_data['total_volume'] = total_volume
        update_data['total_distance'] = total_distance
        update_data['nombre_tours'] = nombre_tours
        update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
        await db.pointages.update_one(
            {"id": existing['id']},
            {"$set": update_data}
        )
        updated = await db.pointages.find_one({"id": existing['id']}, {"_id": 0})
        return updated
    
    # Créer un nouveau pointage
    pointage_data = input.model_dump()
    pointage_data['tours'] = [t.model_dump() if hasattr(t, 'model_dump') else t for t in tours]
    pointage_data['chantier_reference'] = chantier.get('reference')
    pointage_data['chauffeur_nom'] = f"{chauffeur.get('prenom', '')} {chauffeur.get('nom', '')}".strip()
    pointage_data['client_nom'] = chantier.get('client_nom')
    pointage_data['transport_type'] = chantier.get('transport_type')
    pointage_data['avec_gasoil'] = chantier.get('avec_gasoil', True)
    pointage_data['total_volume'] = total_volume
    pointage_data['total_distance'] = total_distance
    pointage_data['nombre_tours'] = nombre_tours
    
    pointage = Pointage(**pointage_data)
    doc = pointage.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.pointages.insert_one(doc)
    return pointage

# Ajouter un tour à un pointage existant
@api_router.post("/pointages/{pointage_id}/tours")
async def add_tour_to_pointage(pointage_id: str, tour: Tour):
    existing = await db.pointages.find_one({"id": pointage_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Pointage non trouvé")
    
    tours = existing.get('tours', [])
    tour_data = tour.model_dump()
    tours.append(tour_data)
    
    # Recalculer les totaux
    total_volume = sum(t.get('volume', 0) for t in tours)
    total_distance = sum(t.get('distance_km', 0) for t in tours)
    
    await db.pointages.update_one(
        {"id": pointage_id},
        {"$set": {
            "tours": tours,
            "total_volume": total_volume,
            "total_distance": total_distance,
            "nombre_tours": len(tours),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.pointages.find_one({"id": pointage_id}, {"_id": 0})
    return updated

# Supprimer un tour d'un pointage
@api_router.delete("/pointages/{pointage_id}/tours/{tour_id}")
async def delete_tour_from_pointage(pointage_id: str, tour_id: str):
    existing = await db.pointages.find_one({"id": pointage_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Pointage non trouvé")
    
    tours = existing.get('tours', [])
    tours = [t for t in tours if t.get('id') != tour_id]
    
    # Recalculer les totaux
    total_volume = sum(t.get('volume', 0) for t in tours)
    total_distance = sum(t.get('distance_km', 0) for t in tours)
    
    await db.pointages.update_one(
        {"id": pointage_id},
        {"$set": {
            "tours": tours,
            "total_volume": total_volume,
            "total_distance": total_distance,
            "nombre_tours": len(tours),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Tour supprimé"}

@api_router.put("/pointages/{pointage_id}", response_model=Pointage)
async def update_pointage(pointage_id: str, input: PointageCreate):
    existing = await db.pointages.find_one({"id": pointage_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Pointage non trouvé")
    
    tours = input.tours or []
    update_data = input.model_dump()
    update_data['tours'] = [t.model_dump() if hasattr(t, 'model_dump') else t for t in tours]
    update_data['total_volume'] = sum(t.volume for t in tours)
    update_data['total_distance'] = sum(t.distance_km for t in tours)
    update_data['nombre_tours'] = len(tours)
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    await db.pointages.update_one({"id": pointage_id}, {"$set": update_data})
    updated = await db.pointages.find_one({"id": pointage_id}, {"_id": 0})
    return updated

@api_router.delete("/pointages/{pointage_id}")
async def delete_pointage(pointage_id: str):
    result = await db.pointages.delete_one({"id": pointage_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pointage non trouvé")
    return {"message": "Pointage supprimé"}

# ============= NOTES DE FRAIS ROUTES =============
@api_router.get("/notes-frais")
async def get_notes_frais(chauffeur_id: Optional[str] = None, statut: Optional[str] = None):
    """Liste des notes de frais"""
    query = {}
    if chauffeur_id:
        query["chauffeur_id"] = chauffeur_id
    if statut:
        query["statut"] = statut
    notes = await db.notes_frais.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return notes

@api_router.get("/notes-frais/{note_id}")
async def get_note_frais(note_id: str):
    """Récupère une note de frais"""
    note = await db.notes_frais.find_one({"id": note_id}, {"_id": 0})
    if not note:
        raise HTTPException(status_code=404, detail="Note de frais non trouvée")
    return note

@api_router.post("/notes-frais")
async def create_note_frais(note: NoteFraisCreate):
    """Créer une note de frais avec photo optionnelle"""
    chauffeur = await db.chauffeurs.find_one({"id": note.chauffeur_id}, {"_id": 0})
    if not chauffeur:
        raise HTTPException(status_code=404, detail="Chauffeur non trouvé")
    
    chantier = None
    if note.chantier_id:
        chantier = await db.chantiers.find_one({"id": note.chantier_id}, {"_id": 0})
    
    # Sauvegarder la photo en base64 si fournie
    photo_url = None
    if note.photo_base64:
        # Stocker la photo en base64 data URL
        photo_url = note.photo_base64 if note.photo_base64.startswith('data:') else f"data:image/jpeg;base64,{note.photo_base64}"
    
    note_dict = {
        "id": str(uuid.uuid4())[:8].upper(),
        "chauffeur_id": note.chauffeur_id,
        "chauffeur_nom": f"{chauffeur.get('prenom', '')} {chauffeur.get('nom', '')}".strip(),
        "pointage_id": note.pointage_id,
        "chantier_id": note.chantier_id,
        "chantier_reference": chantier.get('reference') if chantier else None,
        "date": note.date,
        "montant": note.montant,
        "type_frais": note.type_frais,
        "description": note.description,
        "photo_url": photo_url,
        "statut": "en_attente",
        "date_creation": datetime.now(timezone.utc).isoformat()
    }
    
    await db.notes_frais.insert_one(note_dict)
    # Remove MongoDB _id before returning
    if '_id' in note_dict:
        del note_dict['_id']
    return note_dict

@api_router.put("/notes-frais/{note_id}/statut")
async def update_note_frais_statut(note_id: str, statut: str = Query(..., enum=["en_attente", "valide", "refuse", "rembourse"])):
    """Mettre à jour le statut d'une note de frais"""
    note = await db.notes_frais.find_one({"id": note_id}, {"_id": 0})
    if not note:
        raise HTTPException(status_code=404, detail="Note de frais non trouvée")
    
    await db.notes_frais.update_one({"id": note_id}, {"$set": {"statut": statut}})
    return {"message": f"Statut mis à jour: {statut}"}

@api_router.delete("/notes-frais/{note_id}")
async def delete_note_frais(note_id: str):
    """Supprimer une note de frais"""
    result = await db.notes_frais.delete_one({"id": note_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Note de frais non trouvée")
    return {"message": "Note de frais supprimée"}

# ============= PHOTOS POINTAGE ROUTES =============
@api_router.post("/pointages/{pointage_id}/photos")
async def add_photo_to_pointage(pointage_id: str, photo_base64: str, description: Optional[str] = None):
    """Ajouter une photo à un pointage"""
    pointage = await db.pointages.find_one({"id": pointage_id}, {"_id": 0})
    if not pointage:
        raise HTTPException(status_code=404, detail="Pointage non trouvé")
    
    photos = pointage.get('photos', [])
    
    photo_url = photo_base64 if photo_base64.startswith('data:') else f"data:image/jpeg;base64,{photo_base64}"
    
    photo = {
        "id": str(uuid.uuid4())[:8].upper(),
        "url": photo_url,
        "type": "pointage",
        "description": description,
        "date_creation": datetime.now(timezone.utc).isoformat()
    }
    
    photos.append(photo)
    
    await db.pointages.update_one({"id": pointage_id}, {"$set": {"photos": photos}})
    
    return photo

@api_router.delete("/pointages/{pointage_id}/photos/{photo_id}")
async def delete_photo_from_pointage(pointage_id: str, photo_id: str):
    """Supprimer une photo d'un pointage"""
    pointage = await db.pointages.find_one({"id": pointage_id}, {"_id": 0})
    if not pointage:
        raise HTTPException(status_code=404, detail="Pointage non trouvé")
    
    photos = [p for p in pointage.get('photos', []) if p.get('id') != photo_id]
    
    await db.pointages.update_one({"id": pointage_id}, {"$set": {"photos": photos}})
    
    return {"message": "Photo supprimée"}

# ============= OFFLINE SYNC ROUTES =============
@api_router.post("/sync/pointages")
async def sync_offline_pointages(pointages: List[dict]):
    """Synchroniser les pointages créés hors-ligne"""
    results = {"success": [], "errors": []}
    
    for p in pointages:
        try:
            # Vérifier si le pointage existe déjà
            existing = await db.pointages.find_one({
                "chauffeur_id": p.get("chauffeur_id"),
                "chantier_id": p.get("chantier_id"),
                "date": p.get("date")
            }, {"_id": 0})
            
            if existing:
                # Mettre à jour
                await db.pointages.update_one(
                    {"id": existing["id"]},
                    {"$set": p}
                )
                results["success"].append({"offline_id": p.get("offline_id"), "server_id": existing["id"], "action": "updated"})
            else:
                # Créer
                p["id"] = str(uuid.uuid4())[:8].upper()
                p["date_creation"] = datetime.now(timezone.utc).isoformat()
                await db.pointages.insert_one(p)
                results["success"].append({"offline_id": p.get("offline_id"), "server_id": p["id"], "action": "created"})
        except Exception as e:
            results["errors"].append({"offline_id": p.get("offline_id"), "error": str(e)})
    
    return results

# ============= POINTAGES PDF ROUTES =============
def generate_pointage_pdf_html(pointage: dict, chauffeur: dict, chantier: dict, config: dict) -> str:
    """Génère le HTML pour un PDF de pointage journalier"""
    tours_html = ""
    for i, tour in enumerate(pointage.get('tours', []), 1):
        tours_html += f"""
        <tr>
            <td style="padding: 8px; border: 1px solid #ddd;">{i}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{tour.get('volume', 0)} {tour.get('unite_volume', 'tonnes')}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{tour.get('distance', tour.get('distance_km', 0))} km</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{tour.get('commentaire', '-')}</td>
        </tr>
        """
    
    if not tours_html:
        tours_html = '<tr><td colspan="4" style="padding: 20px; text-align: center; color: #666;">Aucun tour enregistré</td></tr>'
    
    photos_html = ""
    if pointage.get('photos'):
        photos_html = '<div class="photos"><h3>Photos jointes</h3><div class="photo-grid">'
        for photo in pointage.get('photos', []):
            photos_html += f'<img src="{photo.get("url", "")}" alt="Photo" style="max-width: 200px; margin: 5px;"/>'
        photos_html += '</div></div>'
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: Arial, sans-serif; margin: 30px; color: #333; }}
            .header {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 30px; border-bottom: 3px solid #1A4D2E; padding-bottom: 20px; }}
            .company {{ }}
            .company h1 {{ color: #1A4D2E; margin: 0; font-size: 24px; }}
            .doc-info {{ text-align: right; }}
            .doc-title {{ font-size: 20px; font-weight: bold; color: #1A4D2E; }}
            .info-grid {{ display: flex; gap: 30px; margin-bottom: 25px; }}
            .info-box {{ flex: 1; padding: 15px; background: #f5f5f5; border-radius: 8px; border-left: 4px solid #1A4D2E; }}
            .info-box h3 {{ margin: 0 0 10px 0; color: #1A4D2E; font-size: 14px; text-transform: uppercase; }}
            .info-box p {{ margin: 5px 0; }}
            table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            th {{ background: #1A4D2E; color: white; padding: 12px; text-align: left; }}
            .totals {{ background: #f9f9f9; padding: 20px; border-radius: 8px; margin-top: 20px; }}
            .totals-grid {{ display: flex; justify-content: space-around; text-align: center; }}
            .total-item {{ }}
            .total-value {{ font-size: 28px; font-weight: bold; color: #1A4D2E; }}
            .total-label {{ font-size: 12px; color: #666; }}
            .signature {{ margin-top: 40px; display: flex; justify-content: space-between; }}
            .signature-box {{ width: 45%; }}
            .signature-line {{ border-bottom: 1px solid #333; height: 50px; margin-top: 30px; }}
            .photos {{ margin-top: 30px; }}
            .photo-grid {{ display: flex; flex-wrap: wrap; gap: 10px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company">
                <h1>{config.get('raison_sociale', 'TERRE DE BEAUCE')}</h1>
                <p>{config.get('adresse', '')}<br>{config.get('code_postal', '')} {config.get('ville', '')}</p>
            </div>
            <div class="doc-info">
                <div class="doc-title">FICHE DE POINTAGE</div>
                <p><strong>Date:</strong> {pointage.get('date', '')}</p>
                <p><strong>Réf:</strong> PTG-{pointage.get('id', '')}</p>
            </div>
        </div>
        
        <div class="info-grid">
            <div class="info-box">
                <h3>Chauffeur</h3>
                <p><strong>{chauffeur.get('prenom', '')} {chauffeur.get('nom', '')}</strong></p>
                <p>Code: {chauffeur.get('code_acces', '')}</p>
            </div>
            <div class="info-box">
                <h3>Chantier</h3>
                <p><strong>{chantier.get('reference', '')}</strong></p>
                <p>{chantier.get('lieu', '')}</p>
                <p>Client: {chantier.get('client_nom', '')}</p>
            </div>
            <div class="info-box">
                <h3>Type de transport</h3>
                <p><strong>{chantier.get('transport_type', 'solide').capitalize()}</strong></p>
                <p>Gasoil: {'Fourni' if chantier.get('avec_gasoil', True) else 'Non fourni'}</p>
            </div>
        </div>
        
        <h3 style="color: #1A4D2E;">Détail des tours</h3>
        <table>
            <thead>
                <tr>
                    <th style="width: 60px;">Tour</th>
                    <th>Volume</th>
                    <th>Distance</th>
                    <th>Commentaire</th>
                </tr>
            </thead>
            <tbody>
                {tours_html}
            </tbody>
        </table>
        
        <div class="totals">
            <div class="totals-grid">
                <div class="total-item">
                    <div class="total-value">{pointage.get('heures_travaillees', 0)}h</div>
                    <div class="total-label">Heures travaillées</div>
                </div>
                <div class="total-item">
                    <div class="total-value">{pointage.get('nombre_tours', len(pointage.get('tours', [])))}</div>
                    <div class="total-label">Nombre de tours</div>
                </div>
                <div class="total-item">
                    <div class="total-value">{pointage.get('total_volume', 0):.1f}</div>
                    <div class="total-label">Volume total ({chantier.get('transport_type', 'solide') == 'liquide' and 'm³' or 'tonnes'})</div>
                </div>
                <div class="total-item">
                    <div class="total-value">{pointage.get('total_distance', 0):.1f} km</div>
                    <div class="total-label">Distance totale</div>
                </div>
            </div>
        </div>
        
        {photos_html}
        
        <div class="signature">
            <div class="signature-box">
                <p><strong>Signature du chauffeur</strong></p>
                <div class="signature-line"></div>
                <p style="font-size: 12px; color: #666;">Date: {pointage.get('date', '')}</p>
            </div>
            <div class="signature-box">
                <p><strong>Signature du responsable</strong></p>
                <div class="signature-line"></div>
                <p style="font-size: 12px; color: #666;">Date:</p>
            </div>
        </div>
        
        <p style="margin-top: 40px; font-size: 11px; color: #888; text-align: center;">
            Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} - {config.get('raison_sociale', '')}
        </p>
    </body>
    </html>
    """
    return html

def generate_recap_pointages_pdf_html(chantier: dict, pointages: list, config: dict) -> str:
    """Génère le HTML pour un PDF récapitulatif de tous les pointages d'un chantier"""
    
    # Calculer les totaux
    total_heures = sum(p.get('heures_travaillees', 0) for p in pointages)
    total_tours = sum(p.get('nombre_tours', len(p.get('tours', []))) for p in pointages)
    total_volume = sum(p.get('total_volume', 0) for p in pointages)
    total_distance = sum(p.get('total_distance', 0) for p in pointages)
    
    # Générer les lignes du tableau
    rows_html = ""
    for p in sorted(pointages, key=lambda x: x.get('date', '')):
        rows_html += f"""
        <tr>
            <td style="padding: 10px; border: 1px solid #ddd;">{p.get('date', '')}</td>
            <td style="padding: 10px; border: 1px solid #ddd;">{p.get('chauffeur_nom', '')}</td>
            <td style="padding: 10px; border: 1px solid #ddd; text-align: center;">{p.get('heures_travaillees', 0)}h</td>
            <td style="padding: 10px; border: 1px solid #ddd; text-align: center;">{p.get('nombre_tours', len(p.get('tours', [])))}</td>
            <td style="padding: 10px; border: 1px solid #ddd; text-align: right;">{p.get('total_volume', 0):.1f}</td>
            <td style="padding: 10px; border: 1px solid #ddd; text-align: right;">{p.get('total_distance', 0):.1f} km</td>
        </tr>
        """
    
    if not rows_html:
        rows_html = '<tr><td colspan="6" style="padding: 30px; text-align: center; color: #666;">Aucun pointage enregistré</td></tr>'
    
    unite_volume = "m³" if chantier.get('transport_type') == 'liquide' else "tonnes"
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: Arial, sans-serif; margin: 30px; color: #333; }}
            .header {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 30px; border-bottom: 3px solid #1A4D2E; padding-bottom: 20px; }}
            .company h1 {{ color: #1A4D2E; margin: 0; font-size: 24px; }}
            .doc-info {{ text-align: right; }}
            .doc-title {{ font-size: 22px; font-weight: bold; color: #1A4D2E; }}
            .chantier-info {{ background: linear-gradient(135deg, #1A4D2E 0%, #2d7a4a 100%); color: white; padding: 25px; border-radius: 10px; margin-bottom: 30px; }}
            .chantier-info h2 {{ margin: 0 0 10px 0; }}
            .chantier-grid {{ display: flex; gap: 40px; margin-top: 15px; }}
            .chantier-item {{ }}
            .chantier-label {{ font-size: 12px; opacity: 0.8; }}
            .chantier-value {{ font-size: 16px; font-weight: bold; }}
            table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            th {{ background: #1A4D2E; color: white; padding: 12px; text-align: left; }}
            .summary {{ display: flex; gap: 20px; margin-top: 30px; }}
            .summary-box {{ flex: 1; background: #f5f5f5; padding: 20px; border-radius: 10px; text-align: center; border-top: 4px solid #D9A520; }}
            .summary-value {{ font-size: 32px; font-weight: bold; color: #1A4D2E; }}
            .summary-label {{ font-size: 12px; color: #666; margin-top: 5px; }}
            .footer {{ margin-top: 50px; padding-top: 20px; border-top: 1px solid #ddd; }}
            .signature-grid {{ display: flex; justify-content: space-between; margin-top: 30px; }}
            .signature-box {{ width: 30%; text-align: center; }}
            .signature-line {{ border-bottom: 1px solid #333; height: 40px; margin-top: 20px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company">
                <h1>{config.get('raison_sociale', 'TERRE DE BEAUCE')}</h1>
                <p>{config.get('adresse', '')}<br>{config.get('code_postal', '')} {config.get('ville', '')}</p>
                <p>SIRET: {config.get('siret', '')} | TVA: {config.get('tva_intracommunautaire', '')}</p>
            </div>
            <div class="doc-info">
                <div class="doc-title">RÉCAPITULATIF<br>DES POINTAGES</div>
                <p>Document justificatif</p>
            </div>
        </div>
        
        <div class="chantier-info">
            <h2>{chantier.get('reference', '')} - {chantier.get('lieu', '')}</h2>
            <div class="chantier-grid">
                <div class="chantier-item">
                    <div class="chantier-label">Client</div>
                    <div class="chantier-value">{chantier.get('client_nom', '')}</div>
                </div>
                <div class="chantier-item">
                    <div class="chantier-label">Période</div>
                    <div class="chantier-value">{chantier.get('date_debut', '')} → {chantier.get('date_fin', 'En cours')}</div>
                </div>
                <div class="chantier-item">
                    <div class="chantier-label">Type</div>
                    <div class="chantier-value">{chantier.get('transport_type', 'solide').capitalize()} {'(gasoil fourni)' if chantier.get('avec_gasoil', True) else '(sans gasoil)'}</div>
                </div>
                <div class="chantier-item">
                    <div class="chantier-label">Contrat</div>
                    <div class="chantier-value">{chantier.get('numero_contrat', '-')}</div>
                </div>
            </div>
        </div>
        
        <h3 style="color: #1A4D2E;">Détail des pointages ({len(pointages)} jour(s))</h3>
        <table>
            <thead>
                <tr>
                    <th>Date</th>
                    <th>Chauffeur</th>
                    <th style="text-align: center;">Heures</th>
                    <th style="text-align: center;">Tours</th>
                    <th style="text-align: right;">Volume ({unite_volume})</th>
                    <th style="text-align: right;">Distance</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
            <tfoot>
                <tr style="background: #e8f5e9; font-weight: bold;">
                    <td colspan="2" style="padding: 12px; border: 1px solid #ddd;">TOTAUX</td>
                    <td style="padding: 12px; border: 1px solid #ddd; text-align: center;">{total_heures}h</td>
                    <td style="padding: 12px; border: 1px solid #ddd; text-align: center;">{total_tours}</td>
                    <td style="padding: 12px; border: 1px solid #ddd; text-align: right;">{total_volume:.1f}</td>
                    <td style="padding: 12px; border: 1px solid #ddd; text-align: right;">{total_distance:.1f} km</td>
                </tr>
            </tfoot>
        </table>
        
        <div class="summary">
            <div class="summary-box">
                <div class="summary-value">{total_heures}h</div>
                <div class="summary-label">Heures travaillées</div>
            </div>
            <div class="summary-box">
                <div class="summary-value">{total_tours}</div>
                <div class="summary-label">Nombre de tours</div>
            </div>
            <div class="summary-box">
                <div class="summary-value">{total_volume:.1f}</div>
                <div class="summary-label">Volume total ({unite_volume})</div>
            </div>
            <div class="summary-box">
                <div class="summary-value">{total_distance:.1f}</div>
                <div class="summary-label">Distance (km)</div>
            </div>
        </div>
        
        <div class="footer">
            <p style="font-size: 12px; color: #666;">
                Ce document constitue un justificatif des prestations réalisées dans le cadre du chantier référencé ci-dessus.
            </p>
            
            <div class="signature-grid">
                <div class="signature-box">
                    <p><strong>Le prestataire</strong></p>
                    <div class="signature-line"></div>
                    <p style="font-size: 11px;">{config.get('raison_sociale', '')}</p>
                </div>
                <div class="signature-box">
                    <p><strong>Le client</strong></p>
                    <div class="signature-line"></div>
                    <p style="font-size: 11px;">{chantier.get('client_nom', '')}</p>
                </div>
            </div>
        </div>
        
        <p style="margin-top: 40px; font-size: 11px; color: #888; text-align: center;">
            Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} - Réf: RECAP-{chantier.get('id', '')}
        </p>
    </body>
    </html>
    """
    return html

@api_router.get("/pointages/{pointage_id}/pdf")
async def get_pointage_pdf(pointage_id: str):
    """Génère et télécharge le PDF d'un pointage journalier"""
    pointage = await db.pointages.find_one({"id": pointage_id}, {"_id": 0})
    if not pointage:
        raise HTTPException(status_code=404, detail="Pointage non trouvé")
    
    chauffeur = await db.chauffeurs.find_one({"id": pointage.get('chauffeur_id')}, {"_id": 0}) or {}
    chantier = await db.chantiers.find_one({"id": pointage.get('chantier_id')}, {"_id": 0}) or {}
    config = await db.config.find_one({"id": "config_entreprise"}, {"_id": 0}) or {}
    
    html = generate_pointage_pdf_html(pointage, chauffeur, chantier, config)
    
    pdf = HTML(string=html).write_pdf()
    
    filename = f"pointage_{pointage.get('date', '')}_{chauffeur.get('nom', 'chauffeur')}.pdf"
    
    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/chantiers/{chantier_id}/pointages/pdf")
async def get_chantier_pointages_pdf(chantier_id: str):
    """Génère et télécharge le PDF récapitulatif des pointages d'un chantier"""
    chantier = await db.chantiers.find_one({"id": chantier_id}, {"_id": 0})
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier non trouvé")
    
    pointages = await db.pointages.find({"chantier_id": chantier_id}, {"_id": 0}).to_list(1000)
    config = await db.config.find_one({"id": "config_entreprise"}, {"_id": 0}) or {}
    
    html = generate_recap_pointages_pdf_html(chantier, pointages, config)
    
    pdf = HTML(string=html).write_pdf()
    
    # Nom du fichier avec préfixe RH- dérivé de la référence chantier
    chantier_ref_rh = chantier.get('reference', chantier_id)
    if chantier_ref_rh.startswith('CH-'):
        rh_numero = generate_numero_releve_from_chantier(chantier_ref_rh)
    else:
        rh_numero = chantier_ref_rh
    filename = f"{rh_numero}.pdf"
    
    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/factures/{facture_id}/justificatifs")
async def get_facture_justificatifs(facture_id: str):
    """Récupère les justificatifs (pointages) associés à une facture"""
    facture = await db.factures.find_one({"id": facture_id}, {"_id": 0})
    if not facture:
        raise HTTPException(status_code=404, detail="Facture non trouvée")
    
    chantier_id = facture.get('chantier_id')
    pointages = await db.pointages.find({"chantier_id": chantier_id}, {"_id": 0}).to_list(1000)
    chantier = await db.chantiers.find_one({"id": chantier_id}, {"_id": 0}) or {}
    
    return {
        "facture_id": facture_id,
        "facture_numero": facture.get('numero'),
        "chantier_id": chantier_id,
        "chantier_reference": chantier.get('reference'),
        "nombre_pointages": len(pointages),
        "pointages": pointages,
        "totaux": {
            "heures": sum(p.get('heures_travaillees', 0) for p in pointages),
            "tours": sum(p.get('nombre_tours', len(p.get('tours', []))) for p in pointages),
            "volume": sum(p.get('total_volume', 0) for p in pointages),
            "distance": sum(p.get('total_distance', 0) for p in pointages)
        }
    }

@api_router.get("/factures/{facture_id}/justificatifs/pdf")
async def get_facture_justificatifs_pdf(facture_id: str):
    """Génère le PDF des justificatifs pour une facture"""
    facture = await db.factures.find_one({"id": facture_id}, {"_id": 0})
    if not facture:
        raise HTTPException(status_code=404, detail="Facture non trouvée")
    
    chantier_id = facture.get('chantier_id')
    chantier = await db.chantiers.find_one({"id": chantier_id}, {"_id": 0})
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier non trouvé")
    
    pointages = await db.pointages.find({"chantier_id": chantier_id}, {"_id": 0}).to_list(1000)
    config = await db.config.find_one({"id": "config_entreprise"}, {"_id": 0}) or {}
    
    html = generate_recap_pointages_pdf_html(chantier, pointages, config)
    
    pdf = HTML(string=html).write_pdf()
    
    filename = f"justificatifs_facture_{facture.get('numero', facture_id)}.pdf"
    
    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Récapitulatif pointages pour un chantier
@api_router.get("/chantiers/{chantier_id}/recap")
async def get_chantier_recap(chantier_id: str):
    chantier = await db.chantiers.find_one({"id": chantier_id}, {"_id": 0})
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier non trouvé")
    
    pointages = await db.pointages.find({"chantier_id": chantier_id}, {"_id": 0}).to_list(1000)
    
    total_heures = sum(p.get('heures_travaillees', 0) for p in pointages)
    total_tonnage = sum(p.get('tonnage_transporte', 0) for p in pointages)
    total_rotations = sum(p.get('nombre_rotations', 0) for p in pointages)
    nombre_jours = len(set(p.get('date') for p in pointages))
    
    # Calculer le montant selon les tarifs
    montant_ht = 0
    lignes = []
    tarifs = chantier.get('tarifs', [])
    
    for tarif in tarifs:
        methode = tarif.get('methode')
        prix = tarif.get('prix_unitaire', 0)
        
        if methode == 'heure' and total_heures > 0:
            montant = total_heures * prix
            lignes.append({
                "description": f"Heures de travail",
                "quantite": total_heures,
                "unite": "heures",
                "prix_unitaire": prix,
                "montant_ht": montant
            })
            montant_ht += montant
        elif methode == 'tonne' and total_tonnage > 0:
            montant = total_tonnage * prix
            lignes.append({
                "description": f"Transport ({chantier.get('transport_type', 'solide')})",
                "quantite": total_tonnage,
                "unite": "tonnes",
                "prix_unitaire": prix,
                "montant_ht": montant
            })
            montant_ht += montant
        elif methode == 'journee' and nombre_jours > 0:
            montant = nombre_jours * prix
            lignes.append({
                "description": "Forfait journalier",
                "quantite": nombre_jours,
                "unite": "jours",
                "prix_unitaire": prix,
                "montant_ht": montant
            })
            montant_ht += montant
    
    return {
        "chantier": chantier,
        "total_heures": total_heures,
        "total_tonnage": total_tonnage,
        "total_rotations": total_rotations,
        "nombre_jours": nombre_jours,
        "pointages": pointages,
        "lignes_facture": lignes,
        "montant_ht": montant_ht,
        "montant_tva": montant_ht * 0.20,
        "montant_ttc": montant_ht * 1.20
    }

# ============= FACTURES ROUTES =============
@api_router.get("/factures", response_model=List[Facture])
async def get_factures(statut: Optional[FactureStatus] = None, client_id: Optional[str] = None):
    query = {}
    if statut:
        query["statut"] = statut
    if client_id:
        query["client_id"] = client_id
    factures = await db.factures.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return factures

@api_router.get("/factures/{facture_id}", response_model=Facture)
async def get_facture(facture_id: str):
    facture = await db.factures.find_one({"id": facture_id}, {"_id": 0})
    if not facture:
        raise HTTPException(status_code=404, detail="Facture non trouvée")
    return facture

@api_router.post("/factures/generer", response_model=Facture)
async def generer_facture(input: FactureCreate):
    """Génère automatiquement une facture avec logique de facturation complexe:
    - Barèmes kilométriques (solide/liquide × avec/sans gasoil)
    - Règle du minima horaire
    - Lien avec le contrat CCPA
    """
    
    # Récupérer le chantier
    chantier = await db.chantiers.find_one({"id": input.chantier_id}, {"_id": 0})
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier non trouvé")
    
    # Récupérer le client
    client = await db.clients.find_one({"id": chantier.get('client_id')}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    
    # Récupérer la config entreprise
    config = await db.config.find_one({"id": "config_entreprise"}, {"_id": 0})
    if not config:
        config = EntrepriseConfig().model_dump()
    
    # Récupérer les barèmes
    baremes_config = await db.config.find_one({"id": "config_baremes"}, {"_id": 0})
    
    # Récupérer le contrat CCPA associé (s'il existe)
    contrat = await db.contrats_ccpa.find_one({"chantier_id": input.chantier_id}, {"_id": 0})
    
    # Récupérer les pointages
    pointages = await db.pointages.find({"chantier_id": input.chantier_id}, {"_id": 0}).to_list(1000)
    
    if not pointages:
        raise HTTPException(status_code=400, detail="Aucun pointage pour ce chantier")
    
    # Déterminer le type de transport et l'option gasoil
    transport_type = chantier.get('transport_type', 'solide')
    avec_gasoil = chantier.get('avec_gasoil', True)
    
    # Sélectionner le bon barème
    bareme_key = f"{transport_type}_{'avec' if avec_gasoil else 'sans'}_gasoil"
    bareme = baremes_config.get(bareme_key, {}).get('tranches', []) if baremes_config else []
    taux_horaire_minimum = baremes_config.get('taux_horaire_minimum', 0) if baremes_config else 0
    
    # Récupérer le tarif horaire depuis les tarifs du chantier (pour le minima)
    tarifs = chantier.get('tarifs', [])
    tarif_horaire = next((t.get('prix_unitaire', 0) for t in tarifs if t.get('methode') == 'heure'), 0)
    
    # Si pas de tarif horaire dans le chantier, utiliser le taux minimum configuré
    if tarif_horaire == 0:
        tarif_horaire = taux_horaire_minimum
    
    lignes = []
    montant_ht = 0
    
    # Grouper les pointages par jour
    pointages_par_jour = {}
    for p in pointages:
        date = p.get('date')
        if date not in pointages_par_jour:
            pointages_par_jour[date] = []
        pointages_par_jour[date].append(p)
    
    # Traiter chaque jour
    for date, pointages_jour in pointages_par_jour.items():
        heures_jour = sum(p.get('heures_travaillees', 0) for p in pointages_jour)
        
        # Calculer le montant au volume pour cette journée
        montant_volume_jour = 0
        details_volume = []
        
        for pointage in pointages_jour:
            tours = pointage.get('tours', [])
            for tour in tours:
                volume = tour.get('volume', 0)
                distance = tour.get('distance_km', 0)  # Use distance_km field
                
                if volume > 0 and distance > 0 and bareme:
                    # Trouver le prix dans le barème
                    prix_unitaire = 0
                    for tranche in bareme:
                        km_min = tranche.get('km_min', 0)
                        km_max = tranche.get('km_max', float('inf'))
                        if km_min <= distance < km_max:
                            prix_unitaire = tranche.get('prix_tonne_km', 0)
                            break
                    
                    if prix_unitaire > 0:
                        montant_tour = volume * prix_unitaire
                        montant_volume_jour += montant_tour
                        details_volume.append({
                            'volume': volume,
                            'distance': distance,
                            'prix': prix_unitaire,
                            'montant': montant_tour
                        })
        
        # Calculer le montant horaire pour la journée (pour comparaison)
        montant_horaire_jour = heures_jour * tarif_horaire if heures_jour > 0 and tarif_horaire > 0 else 0
        
        # Appliquer la règle du minima horaire
        date_formatted = datetime.strptime(date, "%Y-%m-%d").strftime("%d/%m/%Y")
        
        if montant_volume_jour > 0 and montant_horaire_jour > 0:
            # Comparer les deux modes de facturation
            if montant_volume_jour >= montant_horaire_jour:
                # Facturation au volume (plus avantageuse)
                if details_volume:
                    unite = "m³" if transport_type == "liquide" else "tonnes"
                    total_volume = sum(d['volume'] for d in details_volume)
                    prix_moyen = montant_volume_jour / total_volume if total_volume > 0 else 0
                    lignes.append(LigneFacture(
                        description=f"Transport au volume - {date_formatted} ({len(details_volume)} tour(s))",
                        quantite=round(total_volume, 2),
                        unite=unite,
                        prix_unitaire=round(prix_moyen, 2),
                        montant_ht=round(montant_volume_jour, 2)
                    ))
                    montant_ht += montant_volume_jour
            else:
                # Facturation à l'heure (minima appliqué)
                lignes.append(LigneFacture(
                    description=f"Heures de travail - {date_formatted} (minima horaire appliqué)",
                    quantite=heures_jour,
                    unite="heures",
                    prix_unitaire=tarif_horaire,
                    montant_ht=round(montant_horaire_jour, 2)
                ))
                montant_ht += montant_horaire_jour
        elif montant_volume_jour > 0:
            # Seulement du volume
            if details_volume:
                unite = "m³" if transport_type == "liquide" else "tonnes"
                total_volume = sum(d['volume'] for d in details_volume)
                prix_moyen = montant_volume_jour / total_volume if total_volume > 0 else 0
                lignes.append(LigneFacture(
                    description=f"Transport au volume - {date_formatted}",
                    quantite=round(total_volume, 2),
                    unite=unite,
                    prix_unitaire=round(prix_moyen, 2),
                    montant_ht=round(montant_volume_jour, 2)
                ))
                montant_ht += montant_volume_jour
        elif heures_jour > 0 and tarif_horaire > 0:
            # Seulement des heures
            montant = heures_jour * tarif_horaire
            lignes.append(LigneFacture(
                description=f"Heures de travail - {date_formatted}",
                quantite=heures_jour,
                unite="heures",
                prix_unitaire=tarif_horaire,
                montant_ht=round(montant, 2)
            ))
            montant_ht += montant
    
    # Si toujours pas de lignes, utiliser la facturation simple (tarifs du chantier)
    if not lignes:
        total_heures = sum(p.get('heures_travaillees', 0) for p in pointages)
        total_volume = sum(p.get('total_volume', 0) for p in pointages)
        nombre_jours = len(pointages_par_jour)
        
        for tarif in tarifs:
            methode = tarif.get('methode')
            prix = tarif.get('prix_unitaire', 0)
            
            if methode == 'heure' and total_heures > 0:
                montant = round(total_heures * prix, 2)
                lignes.append(LigneFacture(
                    description=f"Heures de travail - Chantier {chantier.get('reference')}",
                    quantite=total_heures,
                    unite="heures",
                    prix_unitaire=prix,
                    montant_ht=montant
                ))
                montant_ht += montant
            elif methode == 'tonne' and total_volume > 0:
                unite = "m³" if transport_type == "liquide" else "tonnes"
                montant = round(total_volume * prix, 2)
                lignes.append(LigneFacture(
                    description=f"Transport {transport_type} - {chantier.get('lieu')}",
                    quantite=total_volume,
                    unite=unite,
                    prix_unitaire=prix,
                    montant_ht=montant
                ))
                montant_ht += montant
            elif methode == 'journee' and nombre_jours > 0:
                montant = round(nombre_jours * prix, 2)
                lignes.append(LigneFacture(
                    description=f"Forfait journalier - Chantier {chantier.get('reference')}",
                    quantite=nombre_jours,
                    unite="jours",
                    prix_unitaire=prix,
                    montant_ht=montant
                ))
                montant_ht += montant
    
    if not lignes:
        raise HTTPException(status_code=400, detail="Aucune donnée à facturer (vérifiez les pointages, tarifs et barèmes)")
    
    taux_tva = 20.0
    montant_tva = round(montant_ht * (taux_tva / 100), 2)
    montant_ttc = round(montant_ht + montant_tva, 2)
    
    # Construire l'adresse client
    client_adresse = f"{client.get('adresse') or ''}, {client.get('code_postal') or ''} {client.get('ville') or ''}".strip(', ')
    
    # Numéro de facture dérivé de la référence chantier : CH-0003-2026 → FC-0003-2026
    chantier_ref = chantier.get('reference', '')
    if chantier_ref.startswith('CH-'):
        numero_facture = generate_numero_facture_from_chantier(chantier_ref)
    else:
        numero_facture = generate_numero_facture()  # fallback anciens chantiers

    facture = Facture(
        chantier_id=input.chantier_id,
        client_id=client['id'],
        numero=numero_facture,
        date_emission=datetime.now().strftime("%Y-%m-%d"),
        date_echeance=input.date_echeance,
        lignes=[l.model_dump() for l in lignes],
        montant_ht=round(montant_ht, 2),
        taux_tva=taux_tva,
        montant_tva=montant_tva,
        montant_ttc=montant_ttc,
        statut=FactureStatus.BROUILLON,
        notes=input.notes,
        client_raison_sociale=client.get('raison_sociale'),
        client_adresse=client_adresse,
        client_siret=client.get('siret'),
        client_tva=client.get('tva_intracommunautaire'),
        client_email=client.get('email'),
        chantier_reference=chantier.get('reference'),
        chantier_lieu=chantier.get('lieu')
    )
    
    # Ajouter le lien avec le contrat CCPA
    facture_dict = facture.model_dump()
    if contrat:
        facture_dict['contrat_numero'] = contrat.get('numero_contrat')
        facture_dict['contrat_id'] = contrat.get('id')
    
    # Ajouter les informations du compte bancaire
    if input.compte_bancaire_id:
        compte_bancaire = await db.comptes_bancaires.find_one({"id": input.compte_bancaire_id}, {"_id": 0})
        if compte_bancaire:
            facture_dict['compte_bancaire_id'] = compte_bancaire['id']
            facture_dict['compte_bancaire_nom'] = compte_bancaire.get('nom_banque')
            facture_dict['compte_bancaire_iban'] = compte_bancaire.get('iban')
            facture_dict['compte_bancaire_bic'] = compte_bancaire.get('bic')
    else:
        # Utiliser le compte bancaire par défaut
        compte_default = await db.comptes_bancaires.find_one({"is_default": True}, {"_id": 0})
        if compte_default:
            facture_dict['compte_bancaire_id'] = compte_default['id']
            facture_dict['compte_bancaire_nom'] = compte_default.get('nom_banque')
            facture_dict['compte_bancaire_iban'] = compte_default.get('iban')
            facture_dict['compte_bancaire_bic'] = compte_default.get('bic')
        else:
            # Fallback sur la config entreprise (ancien système)
            facture_dict['compte_bancaire_iban'] = config.get('iban')
            facture_dict['compte_bancaire_bic'] = config.get('bic')
    
    facture_dict['created_at'] = facture_dict['created_at'].isoformat()
    await db.factures.insert_one(facture_dict)
    
    # Retourner la facture créée
    created = await db.factures.find_one({"id": facture.id}, {"_id": 0})
    return created

@api_router.put("/factures/{facture_id}/statut")
async def update_facture_statut(facture_id: str, statut: FactureStatus):
    existing = await db.factures.find_one({"id": facture_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Facture non trouvée")
    
    await db.factures.update_one({"id": facture_id}, {"$set": {"statut": statut}})
    updated = await db.factures.find_one({"id": facture_id}, {"_id": 0})
    return updated

@api_router.put("/factures/{facture_id}/compte-bancaire")
async def update_facture_compte_bancaire(facture_id: str, compte_bancaire_id: str):
    """Met à jour le compte bancaire d'une facture"""
    existing = await db.factures.find_one({"id": facture_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Facture non trouvée")
    
    compte_bancaire = await db.comptes_bancaires.find_one({"id": compte_bancaire_id}, {"_id": 0})
    if not compte_bancaire:
        raise HTTPException(status_code=404, detail="Compte bancaire non trouvé")
    
    update_data = {
        "compte_bancaire_id": compte_bancaire['id'],
        "compte_bancaire_nom": compte_bancaire.get('nom_banque'),
        "compte_bancaire_iban": compte_bancaire.get('iban'),
        "compte_bancaire_bic": compte_bancaire.get('bic')
    }
    
    await db.factures.update_one({"id": facture_id}, {"$set": update_data})
    updated = await db.factures.find_one({"id": facture_id}, {"_id": 0})
    return updated

@api_router.delete("/factures/{facture_id}")
async def delete_facture(facture_id: str):
    result = await db.factures.delete_one({"id": facture_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Facture non trouvée")
    return {"message": "Facture supprimée"}

def generate_facture_html(facture: dict, config: dict) -> str:
    """Generate HTML content for an invoice"""
    
    # Format lignes de facture
    lignes_html = ""
    for ligne in facture.get('lignes', []):
        montant = ligne.get('quantite', 0) * ligne.get('prix_unitaire', 0)
        lignes_html += f"""
        <tr>
            <td style="padding: 10px; border-bottom: 1px solid #eee;">{ligne.get('description', '')}</td>
            <td style="padding: 10px; border-bottom: 1px solid #eee; text-align: center;">{ligne.get('quantite', 0)}</td>
            <td style="padding: 10px; border-bottom: 1px solid #eee; text-align: right;">{ligne.get('prix_unitaire', 0):.2f} €</td>
            <td style="padding: 10px; border-bottom: 1px solid #eee; text-align: right;">{montant:.2f} €</td>
        </tr>
        """
    
    if not lignes_html:
        lignes_html = """
        <tr>
            <td colspan="4" style="padding: 20px; text-align: center; color: #999;">Aucune ligne de facturation</td>
        </tr>
        """
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; color: #333; line-height: 1.5; }}
            .header {{ display: flex; justify-content: space-between; margin-bottom: 40px; }}
            .company {{ max-width: 45%; }}
            .company h2 {{ color: #1A4D2E; margin: 0 0 10px 0; font-size: 20px; }}
            .company p {{ margin: 3px 0; font-size: 12px; color: #666; }}
            .invoice-title {{ text-align: right; }}
            .invoice-title h1 {{ color: #1A4D2E; margin: 0; font-size: 28px; }}
            .invoice-title .number {{ font-size: 16px; color: #666; margin-top: 5px; }}
            .invoice-title .date {{ font-size: 12px; color: #999; margin-top: 5px; }}
            .client {{ background: #f9f9f9; padding: 20px; border-radius: 5px; margin-bottom: 30px; }}
            .client-title {{ font-size: 11px; color: #999; text-transform: uppercase; margin-bottom: 10px; }}
            .client h3 {{ margin: 0 0 5px 0; color: #333; font-size: 16px; }}
            .client p {{ margin: 3px 0; font-size: 13px; color: #666; }}
            .chantier {{ background: #e8f5e9; padding: 15px; border-radius: 5px; margin-bottom: 30px; }}
            .chantier-title {{ font-size: 11px; color: #1A4D2E; text-transform: uppercase; margin-bottom: 5px; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; }}
            th {{ background: #1A4D2E; color: white; padding: 12px; text-align: left; font-size: 12px; text-transform: uppercase; }}
            th:nth-child(2), th:nth-child(3), th:nth-child(4) {{ text-align: center; }}
            th:last-child {{ text-align: right; }}
            .totals {{ margin-left: auto; width: 300px; }}
            .total-row {{ display: flex; justify-content: space-between; padding: 8px 0; border-bottom: 1px solid #eee; }}
            .total-row.grand-total {{ border-top: 2px solid #1A4D2E; border-bottom: none; font-size: 18px; font-weight: bold; color: #1A4D2E; padding-top: 15px; }}
            .footer {{ margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; font-size: 11px; color: #999; }}
            .payment {{ background: #fff3cd; padding: 15px; border-radius: 5px; margin-top: 20px; }}
            .payment-title {{ font-weight: bold; color: #856404; margin-bottom: 5px; }}
            .payment p {{ margin: 3px 0; font-size: 12px; color: #856404; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company">
                <h2>{config.get('raison_sociale', 'TERRE DE BEAUCE LOCATION')}</h2>
                <p>{config.get('adresse', '')}</p>
                <p>{config.get('code_postal', '')} {config.get('ville', '')}</p>
                <p>SIRET: {config.get('siret', '')}</p>
                <p>TVA: {config.get('tva_intracommunautaire', '')}</p>
                <p>Email: {config.get('email', '')}</p>
            </div>
            <div class="invoice-title">
                <h1>FACTURE</h1>
                <div class="number">N° {facture.get('numero', '')}</div>
                <div class="date">Date d'émission: {facture.get('date_emission', '')}</div>
            </div>
        </div>
        
        <div class="client">
            <div class="client-title">Facturer à</div>
            <h3>{facture.get('client_raison_sociale', '')}</h3>
            <p>{facture.get('client_adresse', '')}</p>
            <p>SIRET: {facture.get('client_siret', '')}</p>
        </div>
        
        <div class="chantier">
            <div class="chantier-title">Référence chantier</div>
            <strong>{facture.get('chantier_reference', '')}</strong>
            {f"<span style='margin-left: 20px; color: #666;'>Contrat N° {facture.get('contrat_numero')}</span>" if facture.get('contrat_numero') else ""}
        </div>
        
        <table>
            <thead>
                <tr>
                    <th style="width: 50%;">Description</th>
                    <th style="width: 15%;">Quantité</th>
                    <th style="width: 17%;">Prix unitaire</th>
                    <th style="width: 18%;">Montant HT</th>
                </tr>
            </thead>
            <tbody>
                {lignes_html}
            </tbody>
        </table>
        
        <div class="totals">
            <div class="total-row">
                <span>Total HT</span>
                <span>{facture.get('montant_ht', 0):.2f} €</span>
            </div>
            <div class="total-row">
                <span>TVA (20%)</span>
                <span>{facture.get('montant_tva', 0):.2f} €</span>
            </div>
            <div class="total-row grand-total">
                <span>Total TTC</span>
                <span>{facture.get('montant_ttc', 0):.2f} €</span>
            </div>
        </div>
        
        <div class="payment">
            <div class="payment-title">Modalités de paiement</div>
            <p>Paiement à 30 jours à compter de la date d'émission de la facture.</p>
            {f"<p><strong>{facture.get('compte_bancaire_nom', '')}</strong></p>" if facture.get('compte_bancaire_nom') else ""}
            {f"<p>IBAN: {facture.get('compte_bancaire_iban', '')}</p>" if facture.get('compte_bancaire_iban') else (f"<p>IBAN: {config.get('iban', '')}</p>" if config.get('iban') else "")}
            {f"<p>BIC: {facture.get('compte_bancaire_bic', '')}</p>" if facture.get('compte_bancaire_bic') else (f"<p>BIC: {config.get('bic', '')}</p>" if config.get('bic') else "")}
        </div>
        
        <div class="footer">
            <p>{config.get('raison_sociale', '')} - SIRET: {config.get('siret', '')} - TVA: {config.get('tva_intracommunautaire', '')}</p>
            <p>En cas de retard de paiement, une pénalité de 3 fois le taux d'intérêt légal sera appliquée, ainsi qu'une indemnité forfaitaire de 40€ pour frais de recouvrement.</p>
        </div>
    </body>
    </html>
    """
    return html

@api_router.get("/factures/{facture_id}/pdf")
async def get_facture_pdf(facture_id: str):
    """Génère et retourne le PDF de la facture"""
    # Get facture
    facture = await db.factures.find_one({"id": facture_id}, {"_id": 0})
    if not facture:
        raise HTTPException(status_code=404, detail="Facture non trouvée")
    
    # Get entreprise config
    config = await db.config.find_one({"id": "config_entreprise"}, {"_id": 0})
    if not config:
        config = {"raison_sociale": "Terre de Beauce"}
    
    # Generate HTML
    html_content = generate_facture_html(facture, config)
    
    # Generate PDF
    try:
        pdf_bytes = HTML(string=html_content).write_pdf()
        
        filename = f"Facture_{facture['numero'].replace('/', '-')}.pdf"
        
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"inline; filename={filename}"
            }
        )
    except Exception as e:
        logger.error(f"Erreur génération PDF facture: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération du PDF: {str(e)}")

@api_router.get("/factures/{facture_id}/download")
async def download_facture_pdf(facture_id: str):
    """Télécharge le PDF de la facture"""
    # Get facture
    facture = await db.factures.find_one({"id": facture_id}, {"_id": 0})
    if not facture:
        raise HTTPException(status_code=404, detail="Facture non trouvée")
    
    # Get entreprise config
    config = await db.config.find_one({"id": "config_entreprise"}, {"_id": 0})
    if not config:
        config = {"raison_sociale": "Terre de Beauce"}
    
    # Generate HTML
    html_content = generate_facture_html(facture, config)
    
    # Generate PDF
    try:
        pdf_bytes = HTML(string=html_content).write_pdf()
        
        filename = f"Facture_{facture['numero'].replace('/', '-')}.pdf"
        
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except Exception as e:
        logger.error(f"Erreur génération PDF facture: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération du PDF: {str(e)}")

# ============= CONTRATS ROUTES =============
@api_router.get("/contrats", response_model=List[Contrat])
async def get_contrats(client_id: Optional[str] = None, statut: Optional[ContratStatus] = None):
    query = {}
    if client_id:
        query["client_id"] = client_id
    if statut:
        query["statut"] = statut
    contrats = await db.contrats.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return contrats

@api_router.get("/contrats/{contrat_id}", response_model=Contrat)
async def get_contrat(contrat_id: str):
    contrat = await db.contrats.find_one({"id": contrat_id}, {"_id": 0})
    if not contrat:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    return contrat

@api_router.post("/contrats", response_model=Contrat)
async def create_contrat(input: ContratCreate):
    client = await db.clients.find_one({"id": input.client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    
    contrat = Contrat(
        client_id=input.client_id,
        type_transport=input.type_transport,
        reference=generate_reference_contrat(input.type_transport),
        date_creation=datetime.now().strftime("%Y-%m-%d"),
        conditions=input.conditions,
        statut=ContratStatus.BROUILLON,
        client_raison_sociale=client.get('raison_sociale'),
        client_adresse=f"{client.get('adresse')}, {client.get('code_postal')} {client.get('ville')}",
        client_siret=client.get('siret'),
        client_email=client.get('email')
    )
    
    doc = contrat.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.contrats.insert_one(doc)
    return contrat

@api_router.put("/contrats/{contrat_id}/statut")
async def update_contrat_statut(contrat_id: str, statut: ContratStatus):
    existing = await db.contrats.find_one({"id": contrat_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    
    await db.contrats.update_one({"id": contrat_id}, {"$set": {"statut": statut}})
    updated = await db.contrats.find_one({"id": contrat_id}, {"_id": 0})
    return updated

@api_router.delete("/contrats/{contrat_id}")
async def delete_contrat(contrat_id: str):
    result = await db.contrats.delete_one({"id": contrat_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    return {"message": "Contrat supprimé"}

# ============= CONTRATS CCPA ROUTES =============
@api_router.get("/contrats-ccpa")
async def get_contrats_ccpa(chantier_id: Optional[str] = None, client_id: Optional[str] = None, statut: Optional[ContratStatus] = None):
    """Récupère tous les contrats CCPA"""
    query = {}
    if chantier_id:
        query["chantier_id"] = chantier_id
    if client_id:
        query["client_id"] = client_id
    if statut:
        query["statut"] = statut
    contrats = await db.contrats_ccpa.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return contrats

@api_router.get("/contrats-ccpa/{contrat_id}")
async def get_contrat_ccpa(contrat_id: str):
    """Récupère un contrat CCPA par son ID"""
    contrat = await db.contrats_ccpa.find_one({"id": contrat_id}, {"_id": 0})
    if not contrat:
        raise HTTPException(status_code=404, detail="Contrat CCPA non trouvé")
    return contrat

@api_router.get("/chantiers/{chantier_id}/contrat-ccpa")
async def get_contrat_ccpa_by_chantier(chantier_id: str):
    """Récupère le contrat CCPA associé à un chantier"""
    contrat = await db.contrats_ccpa.find_one({"chantier_id": chantier_id}, {"_id": 0})
    if not contrat:
        raise HTTPException(status_code=404, detail="Aucun contrat CCPA pour ce chantier")
    return contrat

@api_router.post("/contrats-ccpa")
async def create_contrat_ccpa(input: ContratCCPACreate):
    """Crée un nouveau contrat CCPA pré-rempli depuis le chantier"""
    # Récupérer le chantier
    chantier = await db.chantiers.find_one({"id": input.chantier_id}, {"_id": 0})
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier non trouvé")
    
    # Vérifier qu'il n'y a pas déjà un contrat pour ce chantier
    existing = await db.contrats_ccpa.find_one({"chantier_id": input.chantier_id}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Un contrat existe déjà pour ce chantier")
    
    # Récupérer le client
    client = await db.clients.find_one({"id": chantier.get('client_id')}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    
    # Déterminer le prix et l'unité depuis les tarifs du chantier
    tarifs = chantier.get('tarifs', [])
    prix_unitaire = 0
    unite_facturation = ""
    if tarifs:
        tarif = tarifs[0]  # Premier tarif
        prix_unitaire = tarif.get('prix_unitaire', 0)
        methode = tarif.get('methode', '')
        if methode == 'heure':
            unite_facturation = "heure effectuée"
        elif methode == 'tonne':
            unite_facturation = "tonne transportée"
        elif methode == 'journee':
            unite_facturation = "journée effectuée"
    
    # Numéro de contrat dérivé de la référence chantier : CH-0003-2026 → CT-0003-2026
    chantier_reference = chantier.get('reference', '')
    if chantier_reference.startswith('CH-'):
        numero_contrat = generate_numero_contrat_from_chantier(chantier_reference)
    else:
        numero_contrat = generate_numero_contrat_ccpa()  # fallback anciens chantiers

    # Créer le contrat CCPA pré-rempli
    contrat = ContratCCPA(
        numero_contrat=numero_contrat,
        chantier_id=input.chantier_id,
        client_id=client['id'],
        client_nom=client.get('raison_sociale') or '',
        client_interlocuteur=client.get('contact_nom') or '',
        client_adresse=f"{client.get('adresse') or ''}, {client.get('code_postal') or ''} {client.get('ville') or ''}",
        client_email=client.get('email') or '',
        client_telephone=client.get('telephone') or '',
        prix_unitaire=prix_unitaire,
        unite_facturation=unite_facturation,
        gasoil_fourni=chantier.get('avec_gasoil', True),
        transport_type=chantier.get('transport_type', 'solide'),
        date_creation=datetime.now().strftime("%Y-%m-%d"),
        statut=ContratStatus.BROUILLON
    )
    
    doc = contrat.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.contrats_ccpa.insert_one(doc)
    
    # Récupérer sans _id
    created = await db.contrats_ccpa.find_one({"id": contrat.id}, {"_id": 0})
    return created

@api_router.put("/contrats-ccpa/{contrat_id}")
async def update_contrat_ccpa(contrat_id: str, input: ContratCCPAUpdate):
    """Met à jour un contrat CCPA (champs modifiables)"""
    existing = await db.contrats_ccpa.find_one({"id": contrat_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Contrat CCPA non trouvé")
    
    # Mettre à jour uniquement les champs fournis
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.contrats_ccpa.update_one({"id": contrat_id}, {"$set": update_data})
    updated = await db.contrats_ccpa.find_one({"id": contrat_id}, {"_id": 0})
    return updated

@api_router.delete("/contrats-ccpa/{contrat_id}")
async def delete_contrat_ccpa(contrat_id: str):
    """Supprime un contrat CCPA"""
    result = await db.contrats_ccpa.delete_one({"id": contrat_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contrat CCPA non trouvé")
    return {"message": "Contrat CCPA supprimé"}

@api_router.get("/contrats-ccpa/{contrat_id}/pdf")
async def get_contrat_ccpa_pdf(contrat_id: str):
    """Génère et retourne le PDF du contrat CCPA"""
    # Get contrat
    contrat = await db.contrats_ccpa.find_one({"id": contrat_id}, {"_id": 0})
    if not contrat:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    
    # Get entreprise config
    config = await db.config.find_one({"id": "config_entreprise"}, {"_id": 0})
    if not config:
        config = {"raison_sociale": "Terre de Beauce"}
    
    # Generate HTML
    html_content = generate_contrat_ccpa_html(contrat, config)
    
    # Generate PDF
    try:
        pdf_bytes = HTML(string=html_content).write_pdf()
        
        filename = f"Contrat_{contrat['numero_contrat'].replace('/', '-')}.pdf"
        
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"inline; filename={filename}"
            }
        )
    except Exception as e:
        logger.error(f"Erreur génération PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération du PDF: {str(e)}")

@api_router.get("/contrats-ccpa/{contrat_id}/download")
async def download_contrat_ccpa_pdf(contrat_id: str):
    """Télécharge le PDF du contrat CCPA"""
    # Get contrat
    contrat = await db.contrats_ccpa.find_one({"id": contrat_id}, {"_id": 0})
    if not contrat:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    
    # Get entreprise config
    config = await db.config.find_one({"id": "config_entreprise"}, {"_id": 0})
    if not config:
        config = {"raison_sociale": "Terre de Beauce"}
    
    # Generate HTML
    html_content = generate_contrat_ccpa_html(contrat, config)
    
    # Generate PDF
    try:
        pdf_bytes = HTML(string=html_content).write_pdf()
        
        filename = f"Contrat_{contrat['numero_contrat'].replace('/', '-')}.pdf"
        
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except Exception as e:
        logger.error(f"Erreur génération PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération du PDF: {str(e)}")

# ============= DASHBOARD STATS =============
@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats():
    total_tracteurs = await db.tracteurs.count_documents({})
    tracteurs_disponibles = await db.tracteurs.count_documents({"statut": VehicleStatus.DISPONIBLE})
    total_equipements = await db.equipements.count_documents({})
    equipements_disponibles = await db.equipements.count_documents({"statut": VehicleStatus.DISPONIBLE})
    total_chauffeurs = await db.chauffeurs.count_documents({})
    chauffeurs_disponibles = await db.chauffeurs.count_documents({"disponible": True})
    total_clients = await db.clients.count_documents({})
    chantiers_en_cours = await db.chantiers.count_documents({"statut": ChantierStatus.EN_COURS})
    chantiers_planifies = await db.chantiers.count_documents({"statut": ChantierStatus.PLANIFIE})
    
    # Factures en attente
    factures_en_attente = await db.factures.count_documents({"statut": {"$in": ["brouillon", "emise", "envoyee"]}})
    
    # CA du mois (factures payées ou signées)
    debut_mois = datetime.now().replace(day=1).strftime("%Y-%m-%d")
    factures_mois = await db.factures.find({
        "statut": {"$in": ["signee", "payee"]},
        "date_emission": {"$gte": debut_mois}
    }, {"_id": 0, "montant_ht": 1}).to_list(1000)
    ca_mois = sum(f.get('montant_ht', 0) for f in factures_mois)
    
    return DashboardStats(
        total_tracteurs=total_tracteurs,
        tracteurs_disponibles=tracteurs_disponibles,
        total_equipements=total_equipements,
        equipements_disponibles=equipements_disponibles,
        total_chauffeurs=total_chauffeurs,
        chauffeurs_disponibles=chauffeurs_disponibles,
        total_clients=total_clients,
        chantiers_en_cours=chantiers_en_cours,
        chantiers_planifies=chantiers_planifies,
        factures_en_attente=factures_en_attente,
        ca_mois=ca_mois
    )

# Root endpoint
@api_router.get("/")
async def root():
    return {"message": "Terre de Beauce ERP API", "version": "2.0.0"}

# ============= DOCUSIGN ROUTES =============

class DocuSignOAuthRequest(BaseModel):
    code: str
    redirect_uri: str

class DocuSignSendRequest(BaseModel):
    document_type: str  # "facture" ou "contrat"
    document_id: str
    signer_email: str
    signer_name: str

# Store DocuSign access token in memory (in production, use database or Redis)
docusign_tokens = {}

@api_router.get("/docusign/auth-url")
async def get_docusign_auth_url(redirect_uri: str):
    """Get DocuSign OAuth authorization URL"""
    if not DOCUSIGN_INTEGRATION_KEY:
        raise HTTPException(status_code=500, detail="DocuSign non configuré")
    
    auth_url = (
        f"https://{DOCUSIGN_AUTH_SERVER}/oauth/auth"
        f"?response_type=code"
        f"&scope=signature"
        f"&client_id={DOCUSIGN_INTEGRATION_KEY}"
        f"&redirect_uri={redirect_uri}"
    )
    return {"auth_url": auth_url}

@api_router.post("/docusign/callback")
async def docusign_oauth_callback(request: DocuSignOAuthRequest):
    """Exchange authorization code for access token"""
    import requests
    
    token_url = f"https://{DOCUSIGN_AUTH_SERVER}/oauth/token"
    
    auth_string = base64.b64encode(
        f"{DOCUSIGN_INTEGRATION_KEY}:{DOCUSIGN_SECRET_KEY}".encode()
    ).decode()
    
    headers = {
        "Authorization": f"Basic {auth_string}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    
    data = {
        "grant_type": "authorization_code",
        "code": request.code,
        "redirect_uri": request.redirect_uri
    }
    
    response = requests.post(token_url, headers=headers, data=data)
    
    if response.status_code != 200:
        raise HTTPException(status_code=400, detail=f"Erreur DocuSign: {response.text}")
    
    token_data = response.json()
    docusign_tokens["access_token"] = token_data.get("access_token")
    docusign_tokens["refresh_token"] = token_data.get("refresh_token")
    docusign_tokens["expires_at"] = datetime.now(timezone.utc).timestamp() + token_data.get("expires_in", 3600)
    
    return {"success": True, "message": "DocuSign connecté avec succès"}

@api_router.get("/docusign/status")
async def get_docusign_status():
    """Check if DocuSign is configured and authenticated"""
    is_configured = bool(DOCUSIGN_INTEGRATION_KEY and DOCUSIGN_SECRET_KEY)
    is_authenticated = bool(docusign_tokens.get("access_token"))
    
    if is_authenticated:
        # Check if token is expired
        expires_at = docusign_tokens.get("expires_at", 0)
        if datetime.now(timezone.utc).timestamp() > expires_at:
            is_authenticated = False
    
    return {
        "configured": is_configured,
        "authenticated": is_authenticated,
        "account_id": DOCUSIGN_ACCOUNT_ID
    }

@api_router.post("/docusign/send-facture/{facture_id}")
async def send_facture_for_signature(facture_id: str, signer_email: str, signer_name: str):
    """Send a facture for electronic signature"""
    
    if not docusign_tokens.get("access_token"):
        raise HTTPException(status_code=401, detail="DocuSign non authentifié. Veuillez vous connecter d'abord.")
    
    # Get facture
    facture = await db.factures.find_one({"id": facture_id}, {"_id": 0})
    if not facture:
        raise HTTPException(status_code=404, detail="Facture non trouvée")
    
    # Get entreprise config
    config = await db.config.find_one({"id": "config_entreprise"}, {"_id": 0})
    if not config:
        config = {"raison_sociale": "Terre de Beauce"}
    
    # Generate simple HTML document for the facture
    html_content = generate_facture_html(facture, config)
    
    try:
        # Create API client
        api_client = ApiClient()
        api_client.host = DOCUSIGN_BASE_URL
        api_client.set_default_header("Authorization", f"Bearer {docusign_tokens['access_token']}")
        
        # Create envelope
        envelope_definition = EnvelopeDefinition(
            email_subject=f"Facture {facture['numero']} - {config.get('raison_sociale')} - Signature requise",
            email_blurb=f"Veuillez signer la facture {facture['numero']} de {config.get('raison_sociale')}.",
            documents=[
                Document(
                    document_base64=base64.b64encode(html_content.encode()).decode(),
                    name=f"Facture_{facture['numero']}.html",
                    file_extension="html",
                    document_id="1"
                )
            ],
            recipients=Recipients(
                signers=[
                    Signer(
                        email=signer_email,
                        name=signer_name,
                        recipient_id="1",
                        routing_order="1",
                        tabs=Tabs(
                            sign_here_tabs=[
                                SignHere(
                                    anchor_string="/sn1/",
                                    anchor_units="pixels",
                                    anchor_x_offset="20",
                                    anchor_y_offset="10"
                                )
                            ]
                        )
                    )
                ]
            ),
            status="sent"
        )
        
        # Send envelope
        envelopes_api = EnvelopesApi(api_client)
        result = envelopes_api.create_envelope(DOCUSIGN_ACCOUNT_ID, envelope_definition=envelope_definition)
        
        # Update facture with envelope ID
        await db.factures.update_one(
            {"id": facture_id},
            {"$set": {
                "docusign_envelope_id": result.envelope_id,
                "statut": "envoyee"
            }}
        )
        
        return {
            "success": True,
            "envelope_id": result.envelope_id,
            "message": f"Facture envoyée à {signer_email} pour signature"
        }
        
    except Exception as e:
        logger.error(f"DocuSign error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur DocuSign: {str(e)}")

@api_router.get("/docusign/envelope/{envelope_id}/status")
async def get_envelope_status(envelope_id: str):
    """Get the status of a DocuSign envelope"""
    
    if not docusign_tokens.get("access_token"):
        raise HTTPException(status_code=401, detail="DocuSign non authentifié")
    
    try:
        api_client = ApiClient()
        api_client.host = DOCUSIGN_BASE_URL
        api_client.set_default_header("Authorization", f"Bearer {docusign_tokens['access_token']}")
        
        envelopes_api = EnvelopesApi(api_client)
        envelope = envelopes_api.get_envelope(DOCUSIGN_ACCOUNT_ID, envelope_id)
        
        return {
            "envelope_id": envelope_id,
            "status": envelope.status,
            "sent_date": envelope.sent_date_time,
            "completed_date": envelope.completed_date_time
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

def generate_facture_html(facture: dict, config: dict) -> str:
    """Generate HTML content for a facture"""
    
    lignes_html = ""
    for ligne in facture.get("lignes", []):
        lignes_html += f"""
        <tr>
            <td style="padding: 8px; border-bottom: 1px solid #eee;">{ligne.get('description', '')}</td>
            <td style="padding: 8px; border-bottom: 1px solid #eee; text-align: right;">{ligne.get('quantite', 0)} {ligne.get('unite', '')}</td>
            <td style="padding: 8px; border-bottom: 1px solid #eee; text-align: right;">{ligne.get('prix_unitaire', 0):.2f} €</td>
            <td style="padding: 8px; border-bottom: 1px solid #eee; text-align: right;">{ligne.get('montant_ht', 0):.2f} €</td>
        </tr>
        """
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; color: #333; }}
            .header {{ display: flex; justify-content: space-between; margin-bottom: 40px; }}
            .company {{ }}
            .company h1 {{ color: #1A4D2E; margin: 0; }}
            .invoice-info {{ text-align: right; }}
            .invoice-number {{ font-size: 24px; font-weight: bold; color: #1A4D2E; }}
            .parties {{ display: flex; justify-content: space-between; margin-bottom: 30px; }}
            .party {{ width: 45%; }}
            .party-title {{ font-size: 12px; color: #666; text-transform: uppercase; margin-bottom: 5px; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; }}
            th {{ background: #1A4D2E; color: white; padding: 10px; text-align: left; }}
            .totals {{ text-align: right; }}
            .totals table {{ width: 300px; margin-left: auto; }}
            .total-row td {{ padding: 8px; }}
            .grand-total {{ font-size: 18px; font-weight: bold; background: #f5f5f5; }}
            .signature {{ margin-top: 50px; padding: 20px; border: 1px dashed #ccc; }}
            .signature-label {{ color: #666; margin-bottom: 10px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="company">
                <h1>{config.get('raison_sociale', 'Terre de Beauce')}</h1>
                <p>{config.get('adresse', '')}<br>
                {config.get('code_postal', '')} {config.get('ville', '')}</p>
                <p>SIRET: {config.get('siret', '')}<br>
                TVA: {config.get('tva_intracommunautaire', '')}</p>
            </div>
            <div class="invoice-info">
                <div class="invoice-number">FACTURE</div>
                <p><strong>N° {facture.get('numero', '')}</strong></p>
                <p>Date: {facture.get('date_emission', '')}<br>
                Échéance: {facture.get('date_echeance', '')}</p>
            </div>
        </div>
        
        <div class="parties">
            <div class="party">
                <div class="party-title">Client</div>
                <p><strong>{facture.get('client_raison_sociale', '')}</strong><br>
                {facture.get('client_adresse', '')}<br>
                {f"SIRET: {facture.get('client_siret', '')}" if facture.get('client_siret') else ""}</p>
            </div>
            <div class="party">
                <div class="party-title">Chantier</div>
                <p><strong>{facture.get('chantier_reference', '')}</strong><br>
                {facture.get('chantier_lieu', '')}</p>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Description</th>
                    <th style="text-align: right;">Quantité</th>
                    <th style="text-align: right;">Prix unitaire</th>
                    <th style="text-align: right;">Total HT</th>
                </tr>
            </thead>
            <tbody>
                {lignes_html}
            </tbody>
        </table>
        
        <div class="totals">
            <table>
                <tr class="total-row">
                    <td>Total HT</td>
                    <td style="text-align: right;">{facture.get('montant_ht', 0):.2f} €</td>
                </tr>
                <tr class="total-row">
                    <td>TVA ({facture.get('taux_tva', 20)}%)</td>
                    <td style="text-align: right;">{facture.get('montant_tva', 0):.2f} €</td>
                </tr>
                <tr class="total-row grand-total">
                    <td>Total TTC</td>
                    <td style="text-align: right;">{facture.get('montant_ttc', 0):.2f} €</td>
                </tr>
            </table>
        </div>
        
        <div class="signature">
            <div class="signature-label">Signature du client (Bon pour accord)</div>
            <p>/sn1/</p>
        </div>
        
        <p style="margin-top: 30px; font-size: 12px; color: #666;">
            {config.get('raison_sociale', '')} - {config.get('email', '')}
        </p>
    </body>
    </html>
    """
    return html

def generate_contrat_ccpa_html(contrat: dict, config: dict) -> str:
    """Generate HTML content for a CCPA contract"""
    
    gasoil_text = "Carburant fourni par la Donneuse d'ordres" if contrat.get('gasoil_fourni', True) else "Carburant non fourni (à charge du Prestataire)"
    transport_type = "solide (céréales, matières premières...)" if contrat.get('transport_type') == 'solide' else "liquide (engrais, produits phytosanitaires...)"
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; color: #333; line-height: 1.6; }}
            .header {{ text-align: center; margin-bottom: 30px; border-bottom: 2px solid #1A4D2E; padding-bottom: 20px; }}
            .header h1 {{ color: #1A4D2E; margin: 0; font-size: 22px; }}
            .header .numero {{ font-size: 14px; color: #666; margin-top: 5px; }}
            .parties {{ display: flex; justify-content: space-between; margin-bottom: 30px; }}
            .party {{ width: 48%; padding: 15px; background: #f9f9f9; border-radius: 5px; }}
            .party-title {{ font-size: 12px; color: #1A4D2E; text-transform: uppercase; font-weight: bold; margin-bottom: 10px; }}
            .article {{ margin-bottom: 20px; }}
            .article h3 {{ color: #1A4D2E; font-size: 14px; margin-bottom: 10px; border-bottom: 1px solid #ddd; padding-bottom: 5px; }}
            .article p {{ margin: 5px 0; font-size: 13px; }}
            .prix-box {{ background: #e8f5e9; padding: 15px; border-radius: 5px; margin: 20px 0; }}
            .prix-box .amount {{ font-size: 20px; font-weight: bold; color: #1A4D2E; }}
            .signature {{ margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; }}
            .signature-grid {{ display: flex; justify-content: space-between; }}
            .signature-box {{ width: 45%; text-align: center; }}
            .signature-label {{ font-size: 12px; color: #666; margin-bottom: 10px; }}
            .signature-line {{ border-bottom: 1px solid #333; height: 60px; margin-bottom: 5px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>CONTRAT CADRE DE PRESTATIONS AGRICOLES</h1>
            <div class="numero">N° {contrat.get('numero_contrat', '')}</div>
        </div>
        
        <div class="parties">
            <div class="party">
                <div class="party-title">Prestataire de services</div>
                <p><strong>{config.get('raison_sociale', 'TERRE DE BEAUCE LOCATION')}</strong></p>
                <p>{config.get('adresse', '')}</p>
                <p>{config.get('code_postal', '')} {config.get('ville', '')}</p>
                <p>SIRET: {config.get('siret', '')}</p>
            </div>
            <div class="party">
                <div class="party-title">Donneuse d'ordres</div>
                <p><strong>{contrat.get('client_nom', '')}</strong></p>
                {f"<p>Contact: {contrat.get('client_interlocuteur')}</p>" if contrat.get('client_interlocuteur') else ""}
                <p>{contrat.get('client_adresse', '')}</p>
                <p>{contrat.get('client_email', '')}</p>
                <p>{contrat.get('client_telephone', '')}</p>
            </div>
        </div>
        
        <div class="article">
            <h3>Article 1 - Objet du contrat</h3>
            <p>Le présent contrat a pour objet de définir les conditions dans lesquelles le Prestataire effectuera des prestations de transport agricole pour le compte de la Donneuse d'ordres.</p>
        </div>
        
        <div class="article">
            <h3>Article 2 - Nature des prestations</h3>
            <p><strong>Type de transport:</strong> Transport de marchandises {transport_type}</p>
            <p><strong>Carburant:</strong> {gasoil_text}</p>
        </div>
        
        <div class="article">
            <h3>Article 3 - Prix</h3>
            <div class="prix-box">
                <p>Toutes les prestations seront facturées au tarif de:</p>
                <p class="amount">{contrat.get('prix_unitaire', 0):.2f} €HT par {contrat.get('unite_facturation', 'unité')}</p>
            </div>
        </div>
        
        <div class="article">
            <h3>Article 4 - Modalités de paiement</h3>
            <p>Les factures sont payables sous 30 jours à compter de la date d'envoi de la facture par mail.</p>
            <p>En cas de retard de paiement, des pénalités seront appliquées conformément à la législation en vigueur.</p>
        </div>
        
        <div class="article">
            <h3>Article 5 - Durée</h3>
            <p>Le présent contrat est conclu pour une durée indéterminée. Il peut être résilié par l'une ou l'autre des parties avec un préavis de 30 jours.</p>
        </div>
        
        <div class="signature">
            <p style="text-align: center; font-size: 12px; color: #666;">Fait en deux exemplaires originaux</p>
            <div class="signature-grid">
                <div class="signature-box">
                    <div class="signature-label">Le Prestataire</div>
                    <div class="signature-line"></div>
                    <p style="font-size: 11px;">{config.get('raison_sociale', '')}</p>
                </div>
                <div class="signature-box">
                    <div class="signature-label">La Donneuse d'ordres (Bon pour accord)</div>
                    <p>/sn1/</p>
                    <p style="font-size: 11px;">{contrat.get('client_nom', '')}</p>
                </div>
            </div>
        </div>
        
        <p style="margin-top: 30px; font-size: 11px; color: #666; text-align: center;">
            {config.get('raison_sociale', '')} - SIRET: {config.get('siret', '')} - {config.get('email', '')}
        </p>
    </body>
    </html>
    """
    return html

@api_router.post("/docusign/send-contrat/{contrat_id}")
async def send_contrat_for_signature(contrat_id: str, signer_email: str, signer_name: str):
    """Send a CCPA contract for electronic signature"""
    
    if not docusign_tokens.get("access_token"):
        raise HTTPException(status_code=401, detail="DocuSign non authentifié. Veuillez vous connecter d'abord.")
    
    # Get contrat
    contrat = await db.contrats_ccpa.find_one({"id": contrat_id}, {"_id": 0})
    if not contrat:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    
    # Get entreprise config
    config = await db.config.find_one({"id": "config_entreprise"}, {"_id": 0})
    if not config:
        config = {"raison_sociale": "Terre de Beauce"}
    
    # Generate HTML document for the contrat
    html_content = generate_contrat_ccpa_html(contrat, config)
    
    try:
        # Create API client
        api_client = ApiClient()
        api_client.host = DOCUSIGN_BASE_URL
        api_client.set_default_header("Authorization", f"Bearer {docusign_tokens['access_token']}")
        
        # Create envelope
        envelope_definition = EnvelopeDefinition(
            email_subject=f"Contrat CCPA {contrat['numero_contrat']} - {config.get('raison_sociale')} - Signature requise",
            email_blurb=f"Veuillez signer le contrat CCPA {contrat['numero_contrat']} de {config.get('raison_sociale')}.",
            documents=[
                Document(
                    document_base64=base64.b64encode(html_content.encode()).decode(),
                    name=f"Contrat_{contrat['numero_contrat']}.html",
                    file_extension="html",
                    document_id="1"
                )
            ],
            recipients=Recipients(
                signers=[
                    Signer(
                        email=signer_email,
                        name=signer_name,
                        recipient_id="1",
                        routing_order="1",
                        tabs=Tabs(
                            sign_here_tabs=[
                                SignHere(
                                    anchor_string="/sn1/",
                                    anchor_units="pixels",
                                    anchor_x_offset="20",
                                    anchor_y_offset="10"
                                )
                            ]
                        )
                    )
                ]
            ),
            status="sent"
        )
        
        # Send envelope
        envelopes_api = EnvelopesApi(api_client)
        result = envelopes_api.create_envelope(DOCUSIGN_ACCOUNT_ID, envelope_definition=envelope_definition)
        
        # Update contrat with envelope ID and status
        await db.contrats_ccpa.update_one(
            {"id": contrat_id},
            {"$set": {
                "docusign_envelope_id": result.envelope_id,
                "docusign_status": "sent",
                "docusign_sent_at": datetime.now(timezone.utc).isoformat(),
                "docusign_signer_email": signer_email,
                "docusign_signer_name": signer_name,
                "statut": "envoye"
            }}
        )
        
        return {
            "success": True,
            "envelope_id": result.envelope_id,
            "message": f"Contrat envoyé à {signer_email} pour signature"
        }
        
    except Exception as e:
        logger.error(f"DocuSign error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur DocuSign: {str(e)}")

@api_router.post("/docusign/sync-status/{document_type}/{document_id}")
async def sync_docusign_status(document_type: str, document_id: str):
    """Synchronize DocuSign status for a facture or contrat"""
    
    if document_type not in ["facture", "contrat"]:
        raise HTTPException(status_code=400, detail="Type de document invalide. Utilisez 'facture' ou 'contrat'")
    
    if not docusign_tokens.get("access_token"):
        raise HTTPException(status_code=401, detail="DocuSign non authentifié")
    
    # Get document
    if document_type == "facture":
        doc = await db.factures.find_one({"id": document_id}, {"_id": 0})
        collection = db.factures
    else:
        doc = await db.contrats_ccpa.find_one({"id": document_id}, {"_id": 0})
        collection = db.contrats_ccpa
    
    if not doc:
        raise HTTPException(status_code=404, detail=f"{document_type.capitalize()} non trouvé")
    
    envelope_id = doc.get("docusign_envelope_id")
    if not envelope_id:
        raise HTTPException(status_code=400, detail="Ce document n'a pas été envoyé pour signature")
    
    try:
        api_client = ApiClient()
        api_client.host = DOCUSIGN_BASE_URL
        api_client.set_default_header("Authorization", f"Bearer {docusign_tokens['access_token']}")
        
        envelopes_api = EnvelopesApi(api_client)
        envelope = envelopes_api.get_envelope(DOCUSIGN_ACCOUNT_ID, envelope_id)
        
        # Map DocuSign status to our status
        docusign_status = envelope.status
        new_statut = doc.get("statut")
        
        if docusign_status == "completed":
            new_statut = "signe" if document_type == "contrat" else "payee"
        elif docusign_status == "declined":
            new_statut = "annule" if document_type == "contrat" else "annulee"
        elif docusign_status == "voided":
            new_statut = "annule" if document_type == "contrat" else "annulee"
        elif docusign_status == "sent":
            new_statut = "envoye" if document_type == "contrat" else "envoyee"
        
        # Update document
        update_data = {
            "docusign_status": docusign_status,
            "statut": new_statut
        }
        
        if envelope.completed_date_time and docusign_status == "completed":
            if document_type == "contrat":
                update_data["date_signature"] = envelope.completed_date_time[:10]
            update_data["docusign_completed_at"] = envelope.completed_date_time
        
        await collection.update_one({"id": document_id}, {"$set": update_data})
        
        updated_doc = await collection.find_one({"id": document_id}, {"_id": 0})
        
        return {
            "envelope_id": envelope_id,
            "docusign_status": docusign_status,
            "document_status": new_statut,
            "sent_date": envelope.sent_date_time,
            "completed_date": envelope.completed_date_time,
            "document": updated_doc
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

@api_router.post("/docusign/sync-all")
async def sync_all_docusign_statuses():
    """Synchronize DocuSign status for all pending documents"""
    
    if not docusign_tokens.get("access_token"):
        raise HTTPException(status_code=401, detail="DocuSign non authentifié")
    
    results = {"factures": [], "contrats": [], "errors": []}
    
    try:
        api_client = ApiClient()
        api_client.host = DOCUSIGN_BASE_URL
        api_client.set_default_header("Authorization", f"Bearer {docusign_tokens['access_token']}")
        envelopes_api = EnvelopesApi(api_client)
        
        # Sync factures
        factures = await db.factures.find(
            {"docusign_envelope_id": {"$exists": True, "$ne": None}, "statut": {"$in": ["envoyee", "en_attente"]}},
            {"_id": 0}
        ).to_list(100)
        
        for facture in factures:
            try:
                envelope = envelopes_api.get_envelope(DOCUSIGN_ACCOUNT_ID, facture["docusign_envelope_id"])
                new_statut = facture["statut"]
                
                if envelope.status == "completed":
                    new_statut = "payee"
                elif envelope.status in ["declined", "voided"]:
                    new_statut = "annulee"
                
                await db.factures.update_one(
                    {"id": facture["id"]},
                    {"$set": {"docusign_status": envelope.status, "statut": new_statut}}
                )
                results["factures"].append({"id": facture["id"], "status": envelope.status})
            except Exception as e:
                results["errors"].append({"type": "facture", "id": facture["id"], "error": str(e)})
        
        # Sync contrats
        contrats = await db.contrats_ccpa.find(
            {"docusign_envelope_id": {"$exists": True, "$ne": None}, "statut": "envoye"},
            {"_id": 0}
        ).to_list(100)
        
        for contrat in contrats:
            try:
                envelope = envelopes_api.get_envelope(DOCUSIGN_ACCOUNT_ID, contrat["docusign_envelope_id"])
                new_statut = contrat["statut"]
                update_data = {"docusign_status": envelope.status}
                
                if envelope.status == "completed":
                    new_statut = "signe"
                    if envelope.completed_date_time:
                        update_data["date_signature"] = envelope.completed_date_time[:10]
                elif envelope.status in ["declined", "voided"]:
                    new_statut = "annule"
                
                update_data["statut"] = new_statut
                await db.contrats_ccpa.update_one({"id": contrat["id"]}, {"$set": update_data})
                results["contrats"].append({"id": contrat["id"], "status": envelope.status})
            except Exception as e:
                results["errors"].append({"type": "contrat", "id": contrat["id"], "error": str(e)})
        
        return results
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# ============= EXPORT ROUTES =============
@api_router.get("/export/factures")
async def export_factures(format: str = Query("csv", enum=["csv", "excel"]), statut: Optional[str] = None):
    """Export des factures en CSV ou Excel"""
    query = {}
    if statut:
        query["statut"] = statut
    
    factures = await db.factures.find(query, {"_id": 0}).sort("date_emission", -1).to_list(1000)
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Factures"
        
        # Header style
        header_fill = PatternFill(start_color="1A4D2E", end_color="1A4D2E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = ["N° Facture", "Date", "Client", "Chantier", "Montant HT", "TVA", "Montant TTC", "Statut", "Contrat CCPA"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, f in enumerate(factures, 2):
            ws.cell(row=row, column=1, value=f.get('numero', ''))
            ws.cell(row=row, column=2, value=f.get('date_emission', ''))
            ws.cell(row=row, column=3, value=f.get('client_raison_sociale', ''))
            ws.cell(row=row, column=4, value=f.get('chantier_reference', ''))
            ws.cell(row=row, column=5, value=f.get('montant_ht', 0))
            ws.cell(row=row, column=6, value=f.get('montant_tva', 0))
            ws.cell(row=row, column=7, value=f.get('montant_ttc', 0))
            ws.cell(row=row, column=8, value=f.get('statut', ''))
            ws.cell(row=row, column=9, value=f.get('contrat_numero', ''))
        
        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=factures_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    else:
        # CSV export
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        
        writer.writerow(["N° Facture", "Date", "Client", "Chantier", "Montant HT", "TVA", "Montant TTC", "Statut", "Contrat CCPA"])
        
        for f in factures:
            writer.writerow([
                f.get('numero', ''),
                f.get('date_emission', ''),
                f.get('client_raison_sociale', ''),
                f.get('chantier_reference', ''),
                f.get('montant_ht', 0),
                f.get('montant_tva', 0),
                f.get('montant_ttc', 0),
                f.get('statut', ''),
                f.get('contrat_numero', '')
            ])
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=factures_{datetime.now().strftime('%Y%m%d')}.csv"}
        )

@api_router.get("/export/pointages")
async def export_pointages(format: str = Query("csv", enum=["csv", "excel"]), chauffeur_id: Optional[str] = None, chantier_id: Optional[str] = None):
    """Export des pointages en CSV ou Excel"""
    query = {}
    if chauffeur_id:
        query["chauffeur_id"] = chauffeur_id
    if chantier_id:
        query["chantier_id"] = chantier_id
    
    pointages = await db.pointages.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Pointages"
        
        header_fill = PatternFill(start_color="1A4D2E", end_color="1A4D2E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ["Date", "Chauffeur", "Chantier", "Heures", "Nb Tours", "Volume Total", "Distance Totale", "Commentaire"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row, p in enumerate(pointages, 2):
            ws.cell(row=row, column=1, value=p.get('date', ''))
            ws.cell(row=row, column=2, value=p.get('chauffeur_nom', ''))
            ws.cell(row=row, column=3, value=p.get('chantier_reference', ''))
            ws.cell(row=row, column=4, value=p.get('heures_travaillees', 0))
            ws.cell(row=row, column=5, value=p.get('nombre_tours', 0))
            ws.cell(row=row, column=6, value=p.get('total_volume', 0))
            ws.cell(row=row, column=7, value=p.get('total_distance', 0))
            ws.cell(row=row, column=8, value=p.get('commentaire', ''))
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=pointages_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    else:
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(["Date", "Chauffeur", "Chantier", "Heures", "Nb Tours", "Volume Total", "Distance Totale", "Commentaire"])
        
        for p in pointages:
            writer.writerow([
                p.get('date', ''),
                p.get('chauffeur_nom', ''),
                p.get('chantier_reference', ''),
                p.get('heures_travaillees', 0),
                p.get('nombre_tours', 0),
                p.get('total_volume', 0),
                p.get('total_distance', 0),
                p.get('commentaire', '')
            ])
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=pointages_{datetime.now().strftime('%Y%m%d')}.csv"}
        )

@api_router.get("/export/chantiers")
async def export_chantiers(format: str = Query("csv", enum=["csv", "excel"]), statut: Optional[str] = None):
    """Export des chantiers en CSV ou Excel"""
    query = {}
    if statut:
        query["statut"] = statut
    
    chantiers = await db.chantiers.find(query, {"_id": 0}).sort("date_debut", -1).to_list(1000)
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Chantiers"
        
        header_fill = PatternFill(start_color="1A4D2E", end_color="1A4D2E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ["Référence", "Client", "Lieu", "Date Début", "Date Fin", "Type Transport", "Gasoil", "Statut"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row, c in enumerate(chantiers, 2):
            ws.cell(row=row, column=1, value=c.get('reference', ''))
            ws.cell(row=row, column=2, value=c.get('client_nom', ''))
            ws.cell(row=row, column=3, value=c.get('lieu', ''))
            ws.cell(row=row, column=4, value=c.get('date_debut', ''))
            ws.cell(row=row, column=5, value=c.get('date_fin', ''))
            ws.cell(row=row, column=6, value=c.get('transport_type', ''))
            ws.cell(row=row, column=7, value="Oui" if c.get('avec_gasoil', True) else "Non")
            ws.cell(row=row, column=8, value=c.get('statut', ''))
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=chantiers_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    else:
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(["Référence", "Client", "Lieu", "Date Début", "Date Fin", "Type Transport", "Gasoil", "Statut"])
        
        for c in chantiers:
            writer.writerow([
                c.get('reference', ''),
                c.get('client_nom', ''),
                c.get('lieu', ''),
                c.get('date_debut', ''),
                c.get('date_fin', ''),
                c.get('transport_type', ''),
                "Oui" if c.get('avec_gasoil', True) else "Non",
                c.get('statut', '')
            ])
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=chantiers_{datetime.now().strftime('%Y%m%d')}.csv"}
        )

# ============= STATISTICS ROUTES =============
@api_router.get("/stats/dashboard")
async def get_dashboard_stats():
    """Statistiques pour le dashboard avancé"""
    now = datetime.now()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start_of_year = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Factures stats
    factures = await db.factures.find({}, {"_id": 0}).to_list(1000)
    factures_mois = [f for f in factures if f.get('date_emission', '') >= start_of_month.strftime('%Y-%m-%d')]
    factures_annee = [f for f in factures if f.get('date_emission', '') >= start_of_year.strftime('%Y-%m-%d')]
    
    ca_mois = sum(f.get('montant_ttc', 0) for f in factures_mois)
    ca_annee = sum(f.get('montant_ttc', 0) for f in factures_annee)
    factures_en_attente = len([f for f in factures if f.get('statut') in ['brouillon', 'emise', 'envoyee']])
    factures_payees = len([f for f in factures if f.get('statut') == 'payee'])
    
    # Chantiers stats
    chantiers = await db.chantiers.find({}, {"_id": 0}).to_list(1000)
    chantiers_actifs = len([c for c in chantiers if c.get('statut') == 'en_cours'])
    chantiers_termines = len([c for c in chantiers if c.get('statut') == 'termine'])
    
    # Pointages stats
    pointages = await db.pointages.find({}, {"_id": 0}).to_list(1000)
    pointages_mois = [p for p in pointages if p.get('date', '') >= start_of_month.strftime('%Y-%m-%d')]
    
    heures_mois = sum(p.get('heures_travaillees', 0) for p in pointages_mois)
    tours_mois = sum(p.get('nombre_tours', 0) for p in pointages_mois)
    volume_mois = sum(p.get('total_volume', 0) for p in pointages_mois)
    
    # Flotte stats
    tracteurs = await db.tracteurs.find({}, {"_id": 0}).to_list(100)
    equipements = await db.equipements.find({}, {"_id": 0}).to_list(100)
    tracteurs_dispo = len([t for t in tracteurs if t.get('statut') == 'disponible'])
    equipements_dispo = len([e for e in equipements if e.get('statut') == 'disponible'])
    
    # Chauffeurs stats
    chauffeurs = await db.chauffeurs.find({}, {"_id": 0}).to_list(100)
    chauffeurs_actifs = len([c for c in chauffeurs if c.get('disponibilite', True)])
    
    # Contrats stats
    contrats = await db.contrats_ccpa.find({}, {"_id": 0}).to_list(1000)
    contrats_signes = len([c for c in contrats if c.get('statut') == 'signe'])
    contrats_en_attente = len([c for c in contrats if c.get('statut') in ['brouillon', 'envoye']])
    
    # Evolution CA par mois (12 derniers mois)
    ca_evolution = []
    for i in range(11, -1, -1):
        month_date = now - timedelta(days=i*30)
        month_start = month_date.replace(day=1).strftime('%Y-%m-%d')
        if month_date.month == 12:
            month_end = month_date.replace(year=month_date.year+1, month=1, day=1).strftime('%Y-%m-%d')
        else:
            month_end = month_date.replace(month=month_date.month+1, day=1).strftime('%Y-%m-%d')
        
        month_factures = [f for f in factures if month_start <= f.get('date_emission', '') < month_end]
        ca_evolution.append({
            'mois': month_date.strftime('%b %Y'),
            'montant': sum(f.get('montant_ttc', 0) for f in month_factures)
        })
    
    # Top clients par CA
    clients_ca = {}
    for f in factures_annee:
        client = f.get('client_raison_sociale', 'Inconnu')
        clients_ca[client] = clients_ca.get(client, 0) + f.get('montant_ttc', 0)
    
    top_clients = sorted([{'nom': k, 'ca': v} for k, v in clients_ca.items()], key=lambda x: x['ca'], reverse=True)[:5]
    
    return {
        'facturation': {
            'ca_mois': round(ca_mois, 2),
            'ca_annee': round(ca_annee, 2),
            'factures_en_attente': factures_en_attente,
            'factures_payees': factures_payees,
            'total_factures': len(factures)
        },
        'chantiers': {
            'actifs': chantiers_actifs,
            'termines': chantiers_termines,
            'total': len(chantiers)
        },
        'activite_mois': {
            'heures': round(heures_mois, 1),
            'tours': tours_mois,
            'volume': round(volume_mois, 1),
            'pointages': len(pointages_mois)
        },
        'flotte': {
            'tracteurs_total': len(tracteurs),
            'tracteurs_dispo': tracteurs_dispo,
            'equipements_total': len(equipements),
            'equipements_dispo': equipements_dispo
        },
        'chauffeurs': {
            'total': len(chauffeurs),
            'actifs': chauffeurs_actifs
        },
        'contrats': {
            'signes': contrats_signes,
            'en_attente': contrats_en_attente,
            'total': len(contrats)
        },
        'evolution_ca': ca_evolution,
        'top_clients': top_clients
    }

# ============= NOTIFICATIONS ROUTES =============
@api_router.get("/notifications")
async def get_notifications():
    """Récupère les notifications et alertes"""
    notifications = []
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')
    
    # Factures en retard de paiement
    factures = await db.factures.find({"statut": {"$in": ["emise", "envoyee", "signee"]}}, {"_id": 0}).to_list(1000)
    for f in factures:
        echeance = f.get('date_echeance', '')
        if echeance and echeance < today:
            jours_retard = (now - datetime.strptime(echeance, '%Y-%m-%d')).days
            notifications.append({
                'type': 'facture_retard',
                'priority': 'high' if jours_retard > 30 else 'medium',
                'title': f"Facture {f.get('numero')} en retard",
                'message': f"Échéance dépassée de {jours_retard} jours - Client: {f.get('client_raison_sociale')}",
                'montant': f.get('montant_ttc'),
                'date': echeance,
                'link': f"/factures"
            })
    
    # Factures arrivant à échéance (7 jours)
    for f in factures:
        echeance = f.get('date_echeance', '')
        if echeance:
            jours_avant = (datetime.strptime(echeance, '%Y-%m-%d') - now).days
            if 0 <= jours_avant <= 7:
                notifications.append({
                    'type': 'facture_echeance',
                    'priority': 'low',
                    'title': f"Facture {f.get('numero')} bientôt à échéance",
                    'message': f"Échéance dans {jours_avant} jours - Client: {f.get('client_raison_sociale')}",
                    'montant': f.get('montant_ttc'),
                    'date': echeance,
                    'link': f"/factures"
                })
    
    # Contrats en attente de signature
    contrats = await db.contrats_ccpa.find({"statut": "envoye"}, {"_id": 0}).to_list(100)
    for c in contrats:
        sent_at = c.get('docusign_sent_at', c.get('date_creation', ''))
        if sent_at:
            try:
                sent_date = datetime.fromisoformat(sent_at.replace('Z', '+00:00')) if 'T' in sent_at else datetime.strptime(sent_at, '%Y-%m-%d')
                jours_attente = (now - sent_date.replace(tzinfo=None)).days
                if jours_attente > 3:
                    notifications.append({
                        'type': 'contrat_attente',
                        'priority': 'medium',
                        'title': f"Contrat {c.get('numero_contrat')} en attente",
                        'message': f"En attente de signature depuis {jours_attente} jours - Client: {c.get('client_nom')}",
                        'date': sent_at[:10] if sent_at else '',
                        'link': f"/contrats"
                    })
            except:
                pass
    
    # Chantiers sans pointage récent (actifs)
    chantiers = await db.chantiers.find({"statut": "en_cours"}, {"_id": 0}).to_list(100)
    for ch in chantiers:
        last_pointage = await db.pointages.find_one(
            {"chantier_id": ch.get('id')},
            {"_id": 0},
            sort=[("date", -1)]
        )
        if last_pointage:
            last_date = last_pointage.get('date', '')
            if last_date:
                jours_sans = (now - datetime.strptime(last_date, '%Y-%m-%d')).days
                if jours_sans > 7:
                    notifications.append({
                        'type': 'chantier_inactif',
                        'priority': 'low',
                        'title': f"Chantier {ch.get('reference')} inactif",
                        'message': f"Aucun pointage depuis {jours_sans} jours",
                        'date': last_date,
                        'link': f"/chantiers"
                    })
        else:
            notifications.append({
                'type': 'chantier_sans_pointage',
                'priority': 'low',
                'title': f"Chantier {ch.get('reference')} sans pointage",
                'message': "Aucun pointage enregistré pour ce chantier actif",
                'link': f"/chantiers"
            })
    
    # Trier par priorité
    priority_order = {'high': 0, 'medium': 1, 'low': 2}
    notifications.sort(key=lambda x: priority_order.get(x.get('priority', 'low'), 2))
    
    return {
        'total': len(notifications),
        'high': len([n for n in notifications if n.get('priority') == 'high']),
        'medium': len([n for n in notifications if n.get('priority') == 'medium']),
        'low': len([n for n in notifications if n.get('priority') == 'low']),
        'notifications': notifications
    }



# ============= DOCUMENTS CHAUFFEUR =============

class DocumentChauffeurBase(BaseModel):
    chauffeur_id: str
    nom: str
    type_document: str = "autre"  # contrat_travail, fiche_paie, commissionnement, autre
    statut: str = "en_attente"    # en_attente, signe, refuse
    docusign_envelope_id: Optional[str] = None
    reference_liee: Optional[str] = None
    notes: Optional[str] = None

class DocumentChauffeurCreate(DocumentChauffeurBase):
    pass

class DocumentChauffeur(DocumentChauffeurBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    signed_at: Optional[datetime] = None

async def get_current_chauffeur_jwt(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Auth chauffeur via JWT pour l'espace documents"""
    if not credentials:
        raise HTTPException(status_code=401, detail="Non authentifié")
    payload = decode_jwt_token(credentials.credentials)
    if not payload or payload.get("type") != "chauffeur":
        raise HTTPException(status_code=401, detail="Token invalide")
    chauffeur = await db.chauffeurs.find_one({"id": payload.get("chauffeur_id")}, {"_id": 0})
    if not chauffeur:
        raise HTTPException(status_code=401, detail="Chauffeur non trouvé")
    return chauffeur

@api_router.get("/chauffeur/documents/en-attente")
async def get_chauffeur_docs_en_attente(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Documents en attente pour le chauffeur connecté - accepte token UUID ou JWT"""
    chauffeur_id = None
    if credentials:
        # Essai JWT
        payload = decode_jwt_token(credentials.credentials)
        if payload and payload.get("type") == "chauffeur":
            chauffeur_id = payload.get("chauffeur_id")
        else:
            # Fallback : token UUID stocké en session
            chauffeur = await db.chauffeurs.find_one({"token_session": credentials.credentials}, {"_id": 0})
            if chauffeur:
                chauffeur_id = chauffeur["id"]
    if not chauffeur_id:
        raise HTTPException(status_code=401, detail="Non authentifié")
    docs = await db.documents_chauffeur.find(
        {"chauffeur_id": chauffeur_id, "statut": "en_attente"}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return docs

@api_router.get("/chauffeur/documents/signes")
async def get_chauffeur_docs_signes(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Documents signés pour le chauffeur connecté"""
    chauffeur_id = None
    if credentials:
        payload = decode_jwt_token(credentials.credentials)
        if payload and payload.get("type") == "chauffeur":
            chauffeur_id = payload.get("chauffeur_id")
        else:
            chauffeur = await db.chauffeurs.find_one({"token_session": credentials.credentials}, {"_id": 0})
            if chauffeur:
                chauffeur_id = chauffeur["id"]
    if not chauffeur_id:
        raise HTTPException(status_code=401, detail="Non authentifié")
    docs = await db.documents_chauffeur.find(
        {"chauffeur_id": chauffeur_id, "statut": "signe"}, {"_id": 0}
    ).sort("signed_at", -1).to_list(100)
    return docs

@api_router.get("/admin/documents-chauffeur")
async def get_all_documents_chauffeur(admin: dict = Depends(get_current_admin)):
    docs = await db.documents_chauffeur.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return docs

@api_router.get("/admin/chauffeurs/{chauffeur_id}/documents")
async def get_documents_par_chauffeur(chauffeur_id: str, admin: dict = Depends(get_current_admin)):
    docs = await db.documents_chauffeur.find(
        {"chauffeur_id": chauffeur_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(200)
    return docs

@api_router.post("/admin/documents-chauffeur", response_model=DocumentChauffeur)
async def create_document_chauffeur(
    input: DocumentChauffeurCreate,
    admin: dict = Depends(get_current_admin)
):
    doc = DocumentChauffeur(**input.model_dump())
    d = doc.model_dump()
    d["created_at"] = d["created_at"].isoformat()
    await db.documents_chauffeur.insert_one(d)
    return doc

@api_router.put("/admin/documents-chauffeur/{doc_id}")
async def update_document_chauffeur(
    doc_id: str,
    statut: Optional[str] = None,
    admin: dict = Depends(get_current_admin)
):
    update = {}
    if statut:
        update["statut"] = statut
        if statut == "signe":
            update["signed_at"] = datetime.now(timezone.utc).isoformat()
    if not update:
        raise HTTPException(status_code=400, detail="Aucune donnée à mettre à jour")
    await db.documents_chauffeur.update_one({"id": doc_id}, {"": update})
    doc = await db.documents_chauffeur.find_one({"id": doc_id}, {"_id": 0})
    return doc

@api_router.delete("/admin/documents-chauffeur/{doc_id}")
async def delete_document_chauffeur(doc_id: str, admin: dict = Depends(get_current_admin)):
    await db.documents_chauffeur.delete_one({"id": doc_id})
    return {"message": "Document supprimé"}

# ============= PORTAIL CLIENT - MODÈLES =============

class ClientPortalLogin(BaseModel):
    identifiant: str
    password: str

class ClientPortalSession(BaseModel):
    client_id: str
    client_nom: str
    token: str

class DocumentStatut(str, Enum):
    EN_ATTENTE = "en_attente"
    SIGNE = "signe"
    REFUSE = "refuse"

class DocumentType(str, Enum):
    CONTRAT = "contrat"
    FACTURE = "facture"
    AUTRE = "autre"

class DocumentClientBase(BaseModel):
    client_id: str
    nom: str
    type_document: DocumentType = DocumentType.AUTRE
    statut: DocumentStatut = DocumentStatut.EN_ATTENTE
    docusign_envelope_id: Optional[str] = None
    reference_liee: Optional[str] = None  # ex: CH-0003-2026
    notes: Optional[str] = None

class DocumentClientCreate(DocumentClientBase):
    pass

class DocumentClient(DocumentClientBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    signed_at: Optional[datetime] = None
    file_url: Optional[str] = None

# ============= PORTAIL CLIENT - AUTH =============

async def get_current_client(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not credentials:
        raise HTTPException(status_code=401, detail="Non authentifié")
    payload = decode_jwt_token(credentials.credentials)
    if not payload or payload.get("type") != "client":
        raise HTTPException(status_code=401, detail="Token invalide")
    client = await db.clients.find_one({"id": payload.get("client_id")}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=401, detail="Client non trouvé")
    return client

@api_router.post("/client/login", response_model=ClientPortalSession)
async def client_login(input: ClientPortalLogin):
    client = await db.clients.find_one({"identifiant_portail": input.identifiant}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=401, detail="Identifiant ou mot de passe incorrect")
    if not client.get("password_portail"):
        raise HTTPException(status_code=401, detail="Accès portail non configuré pour ce client")
    if not verify_password(input.password, client["password_portail"]):
        raise HTTPException(status_code=401, detail="Identifiant ou mot de passe incorrect")
    token = create_jwt_token({"type": "client", "client_id": client["id"]})
    return ClientPortalSession(
        client_id=client["id"],
        client_nom=client.get("raison_sociale", ""),
        token=token
    )

# ============= PORTAIL CLIENT - DOCUMENTS =============

@api_router.get("/client/documents")
async def get_client_documents(current_client: dict = Depends(get_current_client)):
    docs = await db.documents_client.find(
        {"client_id": current_client["id"]}, {"_id": 0}
    ).sort("created_at", -1).to_list(200)
    return docs

@api_router.get("/client/documents/en-attente")
async def get_client_docs_en_attente(current_client: dict = Depends(get_current_client)):
    docs = await db.documents_client.find(
        {"client_id": current_client["id"], "statut": "en_attente"}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return docs

@api_router.get("/client/documents/signes")
async def get_client_docs_signes(current_client: dict = Depends(get_current_client)):
    docs = await db.documents_client.find(
        {"client_id": current_client["id"], "statut": "signe"}, {"_id": 0}
    ).sort("signed_at", -1).to_list(100)
    return docs

@api_router.get("/client/me")
async def get_client_me(current_client: dict = Depends(get_current_client)):
    return {
        "id": current_client["id"],
        "raison_sociale": current_client.get("raison_sociale"),
        "email": current_client.get("email"),
        "ville": current_client.get("ville"),
    }

# ============= ADMIN - GESTION ACCÈS PORTAIL CLIENT =============

@api_router.post("/admin/clients/{client_id}/portail")
async def configurer_acces_portail(
    client_id: str,
    identifiant: str,
    password: str,
    admin: dict = Depends(get_current_admin)
):
    """Configure l'accès portail d'un client (identifiant + mot de passe)"""
    client = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    hashed = hash_password(password)
    await db.clients.update_one(
        {"id": client_id},
        {"": {"identifiant_portail": identifiant, "password_portail": hashed, "portail_actif": True}}
    )
    return {"message": f"Accès portail configuré pour {client.get('raison_sociale')}"}

@api_router.get("/admin/clients/{client_id}/portail")
async def get_portail_status(client_id: str, admin: dict = Depends(get_current_admin)):
    client = await db.clients.find_one({"id": client_id}, {"_id": 0, "password_portail": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    return {
        "portail_actif": client.get("portail_actif", False),
        "identifiant_portail": client.get("identifiant_portail"),
    }

# ============= ADMIN - DOCUMENTS CLIENT =============

@api_router.get("/admin/documents-client")
async def get_all_documents_client(admin: dict = Depends(get_current_admin)):
    docs = await db.documents_client.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return docs

@api_router.post("/admin/documents-client", response_model=DocumentClient)
async def create_document_client(
    input: DocumentClientCreate,
    admin: dict = Depends(get_current_admin)
):
    doc = DocumentClient(**input.model_dump())
    d = doc.model_dump()
    d["created_at"] = d["created_at"].isoformat()
    await db.documents_client.insert_one(d)
    return doc

@api_router.put("/admin/documents-client/{doc_id}")
async def update_document_client(
    doc_id: str,
    statut: Optional[str] = None,
    admin: dict = Depends(get_current_admin)
):
    update = {}
    if statut:
        update["statut"] = statut
        if statut == "signe":
            update["signed_at"] = datetime.now(timezone.utc).isoformat()
    if not update:
        raise HTTPException(status_code=400, detail="Aucune donnée à mettre à jour")
    await db.documents_client.update_one({"id": doc_id}, {"": update})
    doc = await db.documents_client.find_one({"id": doc_id}, {"_id": 0})
    return doc

@api_router.delete("/admin/documents-client/{doc_id}")
async def delete_document_client(doc_id: str, admin: dict = Depends(get_current_admin)):
    await db.documents_client.delete_one({"id": doc_id})
    return {"message": "Document supprimé"}
# ============= TREASURY ROUTES =============

class QuarterConfig(BaseModel):
    id: int
    label: str
    name: str
    start: str
    end: str
    payment_date: str
    color: str = "#E6F1FB"
    text_color: str = "#0C447C"

class QuarterSettingsUpdate(BaseModel):
    quarters: List[QuarterConfig]

class BankAccountTreasuryCreate(BaseModel):
    label: str
    bank: str
    iban: str
    balance: float
    currency: str = "EUR"

class BankAccountTreasury(BankAccountTreasuryCreate):
    id: str
    date_creation: str

class BankAccountTreasuryUpdate(BaseModel):
    label: Optional[str] = None
    bank: Optional[str] = None
    iban: Optional[str] = None
    balance: Optional[float] = None

class AutoDebitCreate(BaseModel):
    account_id: str
    supplier: str
    label: str
    amount: float
    day_of_month: int
    frequency: str
    active: bool = True

class AutoDebitModel(AutoDebitCreate):
    id: str
    date_creation: str

class AutoDebitUpdate(BaseModel):
    supplier: Optional[str] = None
    label: Optional[str] = None
    amount: Optional[float] = None
    day_of_month: Optional[int] = None
    frequency: Optional[str] = None
    active: Optional[bool] = None

class SupplierInvoiceCreate(BaseModel):
    ref: str
    supplier: str
    category: str
    amount: float
    received_date: str
    due_date: str
    quarter_id: int

class SupplierInvoiceModel(SupplierInvoiceCreate):
    id: str
    status: str = "pending"
    date_creation: str

class SupplierInvoiceUpdate(BaseModel):
    ref: Optional[str] = None
    supplier: Optional[str] = None
    category: Optional[str] = None
    amount: Optional[float] = None
    received_date: Optional[str] = None
    due_date: Optional[str] = None
    quarter_id: Optional[int] = None
    status: Optional[str] = None

class CampaignValidate(BaseModel):
    quarter_id: int

class FactorCreate(BaseModel):
    name: str
    short_name: str
    contract_ref: str
    rate: float
    guarantee_rate: float

class FactorModel(FactorCreate):
    id: str
    date_creation: str

class FactorUpdate(BaseModel):
    name: Optional[str] = None
    short_name: Optional[str] = None
    contract_ref: Optional[str] = None
    rate: Optional[float] = None
    guarantee_rate: Optional[float] = None

class FactoringAssignmentCreate(BaseModel):
    factor_id: str
    invoice_ref: str
    client: str
    amount: float
    assigned_date: str
    expected_payment_date: str
    guarantee_release_date: str
    status: str = "pending"

class FactoringAssignmentModel(FactoringAssignmentCreate):
    id: str
    date_creation: str

class FactoringAssignmentUpdate(BaseModel):
    factor_id: Optional[str] = None
    invoice_ref: Optional[str] = None
    client: Optional[str] = None
    amount: Optional[float] = None
    assigned_date: Optional[str] = None
    expected_payment_date: Optional[str] = None
    guarantee_release_date: Optional[str] = None
    status: Optional[str] = None

class ClientInvoiceCreate(BaseModel):
    client: str
    amount: float
    due_date: str
    factored: bool = False
    status: str = "pending"

class ClientInvoiceModel(ClientInvoiceCreate):
    id: str
    date_creation: str

class ClientInvoiceUpdate(BaseModel):
    client: Optional[str] = None
    amount: Optional[float] = None
    due_date: Optional[str] = None
    factored: Optional[bool] = None
    status: Optional[str] = None

class TreasuryWorkerCreate(BaseModel):
    first_name: str
    last_name: str
    role: str
    phone: str
    start_year: int
    active: bool = True

class TreasuryWorkerModel(TreasuryWorkerCreate):
    id: str
    date_creation: str

class TreasuryWorkerUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    role: Optional[str] = None
    phone: Optional[str] = None
    start_year: Optional[int] = None
    active: Optional[bool] = None

class PayslipCreate(BaseModel):
    worker_id: str
    period_start: str
    period_end: str
    gross_salary: float
    net_salary: float
    employer_charges: float
    urssaf: float
    payment_date: str
    status: str = "pending"

class PayslipModel(PayslipCreate):
    id: str
    date_creation: str

class PayslipUpdate(BaseModel):
    period_start: Optional[str] = None
    period_end: Optional[str] = None
    gross_salary: Optional[float] = None
    net_salary: Optional[float] = None
    employer_charges: Optional[float] = None
    urssaf: Optional[float] = None
    payment_date: Optional[str] = None
    status: Optional[str] = None

def treasury_new_id() -> str:
    return str(uuid.uuid4())[:8].upper()

def treasury_now() -> str:
    return datetime.now(timezone.utc).isoformat()

@api_router.get("/treasury/settings/quarters")
async def get_quarter_settings():
    doc = await db.treasury_settings.find_one({"type": "quarters"}, {"_id": 0})
    if not doc:
        return {"quarters": [
            {"id": 1, "label": "T1", "name": "1er trimestre", "start": "01-15", "end": "04-14", "payment_date": "2026-03-31", "color": "#E6F1FB", "text_color": "#0C447C"},
            {"id": 2, "label": "T2", "name": "2ème trimestre", "start": "04-15", "end": "07-14", "payment_date": "2026-06-30", "color": "#EAF3DE", "text_color": "#27500A"},
            {"id": 3, "label": "T3", "name": "3ème trimestre", "start": "07-15", "end": "10-14", "payment_date": "2026-09-30", "color": "#FAEEDA", "text_color": "#633806"},
            {"id": 4, "label": "T4", "name": "4ème trimestre", "start": "10-15", "end": "01-14", "payment_date": "2026-12-31", "color": "#FAECE7", "text_color": "#712B13"},
        ]}
    return doc

@api_router.patch("/treasury/settings/quarters")
async def update_quarter_settings(payload: QuarterSettingsUpdate):
    await db.treasury_settings.update_one(
        {"type": "quarters"},
        {"$set": {"quarters": [q.model_dump() for q in payload.quarters], "updated_at": treasury_now()}},
        upsert=True
    )
    return {"message": "Paramètres trimestres mis à jour"}

@api_router.get("/treasury/bank-accounts")
async def get_treasury_bank_accounts():
    return await db.bank_accounts.find({}, {"_id": 0}).to_list(100)

@api_router.post("/treasury/bank-accounts")
async def create_treasury_bank_account(payload: BankAccountTreasuryCreate):
    doc = payload.model_dump()
    doc["id"] = treasury_new_id()
    doc["date_creation"] = treasury_now()
    await db.bank_accounts.insert_one(doc)
    return doc

@api_router.patch("/treasury/bank-accounts/{account_id}")
async def update_treasury_bank_account(account_id: str, payload: BankAccountTreasuryUpdate):
    existing = await db.bank_accounts.find_one({"id": account_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Compte bancaire non trouvé")
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updates["updated_at"] = treasury_now()
    await db.bank_accounts.update_one({"id": account_id}, {"$set": updates})
    return {"message": "Compte mis à jour"}

@api_router.delete("/treasury/bank-accounts/{account_id}")
async def delete_treasury_bank_account(account_id: str):
    result = await db.bank_accounts.delete_one({"id": account_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Compte bancaire non trouvé")
    return {"message": "Compte supprimé"}

@api_router.get("/treasury/auto-debits")
async def get_auto_debits():
    return await db.auto_debits.find({}, {"_id": 0}).to_list(200)

@api_router.post("/treasury/auto-debits")
async def create_auto_debit(payload: AutoDebitCreate):
    doc = payload.model_dump()
    doc["id"] = treasury_new_id()
    doc["date_creation"] = treasury_now()
    await db.auto_debits.insert_one(doc)
    return doc

@api_router.patch("/treasury/auto-debits/{debit_id}")
async def update_auto_debit(debit_id: str, payload: AutoDebitUpdate):
    existing = await db.auto_debits.find_one({"id": debit_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Prélèvement non trouvé")
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updates["updated_at"] = treasury_now()
    await db.auto_debits.update_one({"id": debit_id}, {"$set": updates})
    return {"message": "Prélèvement mis à jour"}

@api_router.delete("/treasury/auto-debits/{debit_id}")
async def delete_auto_debit(debit_id: str):
    result = await db.auto_debits.delete_one({"id": debit_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Prélèvement non trouvé")
    return {"message": "Prélèvement supprimé"}

@api_router.get("/treasury/supplier-invoices")
async def get_supplier_invoices(quarter_id: Optional[int] = None, status: Optional[str] = None):
    query = {}
    if quarter_id:
        query["quarter_id"] = quarter_id
    if status:
        query["status"] = status
    return await db.supplier_invoices.find(query, {"_id": 0}).to_list(500)

@api_router.post("/treasury/supplier-invoices")
async def create_supplier_invoice(payload: SupplierInvoiceCreate):
    doc = payload.model_dump()
    doc["id"] = treasury_new_id()
    doc["status"] = "pending"
    doc["date_creation"] = treasury_now()
    settings = await db.treasury_settings.find_one({"type": "quarters"}, {"_id": 0})
    if settings and "quarters" in settings:
        received = datetime.fromisoformat(doc["received_date"])
        for q in settings["quarters"]:
            mm, dd = q["end"].split("-")
            end_date = datetime(received.year, int(mm), int(dd))
            if end_date < received:
                end_date = datetime(received.year + 1, int(mm), int(dd))
            if 0 <= (end_date - received).days < 30:
                doc["quarter_id"] = (q["id"] % 4) + 1
                break
    await db.supplier_invoices.insert_one(doc)
    return doc

@api_router.patch("/treasury/supplier-invoices/{invoice_id}")
async def update_supplier_invoice(invoice_id: str, payload: SupplierInvoiceUpdate):
    existing = await db.supplier_invoices.find_one({"id": invoice_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Facture fournisseur non trouvée")
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updates["updated_at"] = treasury_now()
    await db.supplier_invoices.update_one({"id": invoice_id}, {"$set": updates})
    return {"message": "Facture mise à jour"}

@api_router.delete("/treasury/supplier-invoices/{invoice_id}")
async def delete_supplier_invoice(invoice_id: str):
    result = await db.supplier_invoices.delete_one({"id": invoice_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Facture fournisseur non trouvée")
    return {"message": "Facture supprimée"}

@api_router.post("/treasury/payment-campaigns/validate")
async def validate_payment_campaign(payload: CampaignValidate):
    result = await db.supplier_invoices.update_many(
        {"quarter_id": payload.quarter_id, "status": "pending"},
        {"$set": {"status": "paid", "paid_at": treasury_now()}}
    )
    await db.payment_campaigns.update_one(
        {"quarter_id": payload.quarter_id},
        {"$set": {"status": "validated", "validated_at": treasury_now()}},
        upsert=True
    )
    return {"message": "Campagne validée", "invoices_updated": result.modified_count}

@api_router.get("/treasury/factors")
async def get_factors():
    return await db.factors.find({}, {"_id": 0}).to_list(50)

@api_router.post("/treasury/factors")
async def create_factor(payload: FactorCreate):
    doc = payload.model_dump()
    doc["id"] = treasury_new_id()
    doc["date_creation"] = treasury_now()
    await db.factors.insert_one(doc)
    return doc

@api_router.patch("/treasury/factors/{factor_id}")
async def update_factor(factor_id: str, payload: FactorUpdate):
    existing = await db.factors.find_one({"id": factor_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Factor non trouvé")
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updates["updated_at"] = treasury_now()
    await db.factors.update_one({"id": factor_id}, {"$set": updates})
    return {"message": "Factor mis à jour"}

@api_router.delete("/treasury/factors/{factor_id}")
async def delete_factor(factor_id: str):
    result = await db.factors.delete_one({"id": factor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Factor non trouvé")
    return {"message": "Factor supprimé"}

@api_router.get("/treasury/factoring-assignments")
async def get_factoring_assignments(factor_id: Optional[str] = None, status: Optional[str] = None):
    query = {}
    if factor_id:
        query["factor_id"] = factor_id
    if status:
        query["status"] = status
    return await db.factoring_assignments.find(query, {"_id": 0}).to_list(500)

@api_router.post("/treasury/factoring-assignments")
async def create_factoring_assignment(payload: FactoringAssignmentCreate):
    doc = payload.model_dump()
    doc["id"] = treasury_new_id()
    doc["date_creation"] = treasury_now()
    await db.factoring_assignments.insert_one(doc)
    return doc

@api_router.patch("/treasury/factoring-assignments/{assignment_id}")
async def update_factoring_assignment(assignment_id: str, payload: FactoringAssignmentUpdate):
    existing = await db.factoring_assignments.find_one({"id": assignment_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Cession non trouvée")
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updates["updated_at"] = treasury_now()
    await db.factoring_assignments.update_one({"id": assignment_id}, {"$set": updates})
    return {"message": "Cession mise à jour"}

@api_router.delete("/treasury/factoring-assignments/{assignment_id}")
async def delete_factoring_assignment(assignment_id: str):
    result = await db.factoring_assignments.delete_one({"id": assignment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Cession non trouvée")
    return {"message": "Cession supprimée"}

@api_router.get("/treasury/client-invoices")
async def get_client_invoices_treasury(status: Optional[str] = None):
    query = {}
    if status:
        query["status"] = status
    return await db.client_invoices.find(query, {"_id": 0}).to_list(500)

@api_router.post("/treasury/client-invoices")
async def create_client_invoice_treasury(payload: ClientInvoiceCreate):
    doc = payload.model_dump()
    doc["id"] = treasury_new_id()
    doc["date_creation"] = treasury_now()
    await db.client_invoices.insert_one(doc)
    return doc

@api_router.patch("/treasury/client-invoices/{invoice_id}")
async def update_client_invoice_treasury(invoice_id: str, payload: ClientInvoiceUpdate):
    existing = await db.client_invoices.find_one({"id": invoice_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Facture client non trouvée")
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updates["updated_at"] = treasury_now()
    await db.client_invoices.update_one({"id": invoice_id}, {"$set": updates})
    return {"message": "Facture cliente mise à jour"}

@api_router.delete("/treasury/client-invoices/{invoice_id}")
async def delete_client_invoice_treasury(invoice_id: str):
    result = await db.client_invoices.delete_one({"id": invoice_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Facture client non trouvée")
    return {"message": "Facture cliente supprimée"}

@api_router.get("/treasury/workers")
async def get_treasury_workers():
    return await db.treasury_workers.find({}, {"_id": 0}).to_list(100)

@api_router.post("/treasury/workers")
async def create_treasury_worker(payload: TreasuryWorkerCreate):
    doc = payload.model_dump()
    doc["id"] = treasury_new_id()
    doc["date_creation"] = treasury_now()
    await db.treasury_workers.insert_one(doc)
    return doc

@api_router.patch("/treasury/workers/{worker_id}")
async def update_treasury_worker(worker_id: str, payload: TreasuryWorkerUpdate):
    existing = await db.treasury_workers.find_one({"id": worker_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Saisonnier non trouvé")
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updates["updated_at"] = treasury_now()
    await db.treasury_workers.update_one({"id": worker_id}, {"$set": updates})
    return {"message": "Saisonnier mis à jour"}

@api_router.delete("/treasury/workers/{worker_id}")
async def delete_treasury_worker(worker_id: str):
    result = await db.treasury_workers.delete_one({"id": worker_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Saisonnier non trouvé")
    return {"message": "Saisonnier supprimé"}

@api_router.get("/treasury/payslips")
async def get_payslips(worker_id: Optional[str] = None, status: Optional[str] = None):
    query = {}
    if worker_id:
        query["worker_id"] = worker_id
    if status:
        query["status"] = status
    return await db.payslips.find(query, {"_id": 0}).to_list(500)

@api_router.post("/treasury/payslips")
async def create_payslip(payload: PayslipCreate):
    doc = payload.model_dump()
    doc["id"] = treasury_new_id()
    doc["date_creation"] = treasury_now()
    await db.payslips.insert_one(doc)
    return doc

@api_router.patch("/treasury/payslips/{payslip_id}")
async def update_payslip(payslip_id: str, payload: PayslipUpdate):
    existing = await db.payslips.find_one({"id": payslip_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Fiche de paie non trouvée")
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updates["updated_at"] = treasury_now()
    await db.payslips.update_one({"id": payslip_id}, {"$set": updates})
    return {"message": "Fiche de paie mise à jour"}

@api_router.delete("/treasury/payslips/{payslip_id}")
async def delete_payslip(payslip_id: str):
    result = await db.payslips.delete_one({"id": payslip_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fiche de paie non trouvée")
    return {"message": "Fiche de paie supprimée"}

@api_router.get("/treasury/dashboard")
async def get_treasury_dashboard():
    accounts = await db.bank_accounts.find({}, {"_id": 0}).to_list(100)
    auto_debits = await db.auto_debits.find({"active": True}, {"_id": 0}).to_list(200)
    supplier_pending = await db.supplier_invoices.find({"status": "pending"}, {"_id": 0}).to_list(500)
    client_pending = await db.client_invoices.find({"status": "pending"}, {"_id": 0}).to_list(500)
    factors = await db.factors.find({}, {"_id": 0}).to_list(50)
    factoring_pending = await db.factoring_assignments.find({"status": "pending"}, {"_id": 0}).to_list(200)
    campaigns = await db.payment_campaigns.find({}, {"_id": 0}).to_list(20)
    payslips_pending = await db.payslips.find({"status": "pending"}, {"_id": 0}).to_list(200)
    factors_map = {f["id"]: f for f in factors}
    total_balance = sum(a["balance"] for a in accounts)
    total_guarantee = sum(
        a["amount"] * factors_map.get(a["factor_id"], {}).get("guarantee_rate", 0) / 100
        for a in factoring_pending
    )
    monthly_auto = sum(
        d["amount"] if d["frequency"] == "monthly"
        else d["amount"] / 3 if d["frequency"] == "quarterly"
        else d["amount"] / 12
        for d in auto_debits
    )
    quarter_totals = {}
    for inv in supplier_pending:
        qid = str(inv.get("quarter_id"))
        quarter_totals[qid] = quarter_totals.get(qid, 0) + inv["amount"]
    return {
        "total_balance": total_balance,
        "available_after_guarantee": total_balance - total_guarantee,
        "total_guarantee_immobilized": total_guarantee,
        "total_client_pending": sum(i["amount"] for i in client_pending),
        "total_supplier_pending": sum(i["amount"] for i in supplier_pending),
        "monthly_auto_debits": monthly_auto,
        "total_payroll_pending": sum(p["net_salary"] + p["employer_charges"] for p in payslips_pending),
        "total_urssaf_pending": sum(p["urssaf"] for p in payslips_pending),
        "quarter_totals": quarter_totals,
        "accounts": accounts,
        "campaigns": campaigns,
    }
# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
